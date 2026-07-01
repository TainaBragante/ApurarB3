import io
import os
import re
import tkinter as tk
import openpyxl
import traceback
import fitz
import sys
import json
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from copy import deepcopy
from datetime import date
from typing import Callable, Dict, List, Optional, Tuple
from tkinter import filedialog, messagebox
from PIL import Image, ImageTk
from correpy.parsers.brokerage_notes.parser_factory import ParserFactory
from correpy.parsers.exceptions import InvalidPasswordException
import pdfplumber


###############################################################################
# CONSTANTES DE TIPO DE ATIVO
###############################################################################

# Identificadores canônicos de tipo de ativo usados internamente.
# Esses valores são gravados no campo `asset_type` de cada Operation.
ASSET_TYPE_ON     = "ON"      # Ação ordinária                → tem isenção, 15% swing
ASSET_TYPE_PN     = "PN"      # Ação preferencial             → tem isenção, 15% swing
ASSET_TYPE_FII    = "FII"     # Fundo Imobiliário             → sem isenção, 20% swing
ASSET_TYPE_FIAGRO = "FIAGRO"  # Fundo Agroindustrial          → sem isenção, 20% swing
ASSET_TYPE_BDR    = "BDR"     # Brazilian Depositary Receipt  → sem isenção, 15% swing
ASSET_TYPE_UNITS  = "UNITS"   # Units                         → sem isenção, 15% swing
ASSET_TYPE_ETF    = "ETF"     # Exchange Traded Fund          → sem isenção, 15% swing

# Conjunto de tipos que NÃO têm isenção de 20k
ASSET_TYPES_SEM_ISENCAO = {
    ASSET_TYPE_FII, ASSET_TYPE_FIAGRO,
    ASSET_TYPE_BDR, ASSET_TYPE_UNITS, ASSET_TYPE_ETF,
}

# Conjunto de tipos que TÊM isenção de 20k (ON e PN)
ASSET_TYPES_COM_ISENCAO = {ASSET_TYPE_ON, ASSET_TYPE_PN}

# Limite de isenção mensal de IR para ON e PN (R$ 20.000,00 em vendas)
LIMITE_ISENCAO_ON_PN = Decimal("20000.00")

# Lista de 112 tickers de ETFs conhecidos para detecção automática (sem necessidade de perguntar ao usuário).
# Lista exportada do site da B3 "https://www.b3.com.br/pt_br/produtos-e-servicos/negociacao/renda-variavel/etf/renda-variavel/etfs-listados/" 
ETF_TICKERS_CONHECIDOS = {
    "ACWI11", "AGRI11", "ALUG11", "ARGE11", "AURO11", "AUVP11", "B3BR11", "BBOI11", "BBOV11", "BBSD11", "BCIC11", "BDEF11", "BDOM11", "BEST11", "BIZD11", "BMMT11", "BOVA11", "BOVB11", "BOVS11", "BOVV11", "BOVX11", "BRAX11", "BRAZ11", "BREW11", "BRXC11", "BULZ11", "BVBR11", "BXPO11", "CAPE11", "CASA11", "CHIP11", "CMDB11", "COIN11", "CORN11", "DIVD11", "DIVO11", "DOLA11", "DOLB11", "DOLX11", "DVER11", "ECOO11", "ELAS11", "ESGB11", "EWBZ11", "FIND11", "FIXX11", "GDIV11", "GENB11", "GLDI11", "GLDX11", "GOLB11", "GOLD11", "GOLX11", "GOVE11", "GPUS11", "GXUS11", "HIGH11", "HTEK11", "IBOB11", "ISUS11", "IVVB11", "IVWO11", "IWMI11", "JOGO11", "LVOL11", "MATB11", "MILL11", "NASD11", "NBOV11", "NDIV11", "NSDV11", "NUCL11", "OURO11", "PEVC11", "PIBB11", "PIPE11", "PKIN11", "QLBR11", "QQQI11", "QQQQ11", "REVE11", "RICO11", "SCVB11", "SILK11", "SLVR11", "SMAB11", "SMAC11", "SMAL11", "SPBZ11", "SPUB11", "SPVT11", "SPXB11", "SPXH11", "SPXI11", "SPXR11", "SPYI11", "SPYR11", "SVAL11", "TECK11", "TECX11", "TIRB11", "TRIG11", "USAL11", "USTK11", "UTEC11", "UTLL11", "VWRA11", "WRLD11", "XBCI11", "XBOV11", "XINA11", "XSPI11",
}

# Lista de 13 tickers de UNITS conhecidos para detecção automática (sem necessidade de perguntar ao usuário).
# Lista exportada do site da B3 "https://www.b3.com.br/main.jsp?lumPageId=8A6A8C244DFC31D8014DFC8051B81CFD&lumA=1&lumII=8A80CB81633FBF0B016340E824360853&locale=en_US&doui_processActionId=setLocaleProcessAction&doui_action=setLocaleProcessAction&doui_actionId=setLocaleProcessAction&doui_processActionId=setLocaleProcessAction&doui_processActionId=setLocaleProcessAction" 
UNITS_TICKERS_CONHECIDOS = {
    "ALUP11", "BRBI11", "BPAC11", "CPLE11", "ENGI11", "KLBN11", "PPLA11", "SAPR11", "SANB11", "SULA11", "TAEE11", "IGTI11", "RBNS11"
}

# Fragmentos de CNPJ para detecção de corretoras não-CorrePy
# Mapeamento: fragmento sem formatação → identificador interno 
BROKER_CNPJ_MAP = {
    "18945670": "inter",     # Inter DTVM Ltda. - CNPJ 18.945.670/0001-46
    "27652684": "genial",    # Genial CCTVM S/A - CNPJ 27.652.684/0001-62
    "05816451": "genial",    # Genial CCTVM S/A - CNPJ 05.816.451/0001-15
}


###############################################################################
# ESTRUTURA DE DADOS PARA OPERAÇÕES
###############################################################################

@dataclass
class Operation:
    """Representa uma operação individual extraída da nota de corretagem.

    Atributos principais:
    - ref_date         : data do pregão.
    - ticker           : código do ativo (ex.: PETR4, HGLG11).
    - name             : nome/descrição completa do ativo como aparece na nota.
    - transaction_type : "buy" ou "sell".
    - amount           : quantidade negociada.
    - unit_price       : preço unitário.
    - total_value      : valor financeiro total (amount x unit_price).
    - allocated_fee    : taxas rateadas proporcionalmente atribuídas a esta operação.
    - irrf             : IR retido na fonte indicado na nota.
    - note_file        : nome do arquivo PDF de origem.
    - category         : "swing", "day" ou "fii" — preenchido por classify_operations().
    - asset_type       : tipo do ativo (ON, PN, FII, FIAGRO, BDR, UNITS, ETF) — preenchido por detect_asset_type() ou solicitado ao usuário.
    - parser_used      : qual camada de parser leu esta operação ("correpy" ou "pdfplumber_<corretora>").
    """

    ref_date: date
    ticker: Optional[str]
    name: str
    transaction_type: str
    amount: Decimal
    unit_price: Decimal
    total_value: Decimal
    allocated_fee: Decimal
    irrf: Decimal
    note_file: str
    category: Optional[str] = field(default=None)
    asset_type: Optional[str] = field(default=None)   
    parser_used: str = field(default="correpy")


###############################################################################
# CAMADA 1 — CORREPY (parser principal via biblioteca)
###############################################################################

# Estrutura normalizada para as taxas do Resumo Financeiro das notas SINACOR.
@dataclass
class FinancialSummary:
    settlement_fee: Decimal = Decimal("0")   # Taxa de liquidação
    registration_fee: Decimal = Decimal("0") # Taxa de registro
    transfer_fee: Decimal = Decimal("0")     # Taxa de transferência de ativos
    ana_fee: Decimal = Decimal("0")          # Taxa ANA
    emoluments: Decimal = Decimal("0")       # Emolumentos
    operational_fee: Decimal = Decimal("0")  # Taxa operacional
    execution_fee: Decimal = Decimal("0")    # Execução
    custody_fee: Decimal = Decimal("0")      # Taxa de custódia
    taxes: Decimal = Decimal("0")            # Impostos
    other_fee: Decimal = Decimal("0")        # Outras
    irrf: Decimal = Decimal("0")             # IRRF

    source: str = ""
    warnings: list = field(default_factory=list)

    def total_allocatable_fees(self, include_transfer_fee: bool = True) -> Decimal:
        """Soma apenas taxas que entram no rateio das operações."""
        base = (
            self.settlement_fee + self.registration_fee +
            self.ana_fee + self.emoluments + self.operational_fee +
            self.execution_fee + self.custody_fee +
            self.taxes + self.other_fee
        )
        if include_transfer_fee:
            base += self.transfer_fee
        return base

    def has_any_fee(self) -> bool:
        """Retorna True se pelo menos um campo de taxa foi extraído com valor não-zero."""
        return self.total_allocatable_fees() > Decimal("0") or self.irrf > Decimal("0")


# Padrões centralizados dos rótulos de taxas aceitos no Resumo Financeiro.
FEE_LABEL_PATTERNS = {
    "settlement_fee": [
        r"Taxa\s+de\s+liquida[cç][aã]o",
        r"Taxa\s+de\s+liquida[cç][aã]o\s*/\s*CCP",
    ],
    "registration_fee": [
        r"Taxa\s+de\s+Registro",
    ],
    "transfer_fee": [
        r"Taxa\s+de\s+(?:Transfer[eê]ncia|Tranfer[eê]ncia|Transf\.?)\s+de\s+Ativos",
    ],
    "emoluments": [
        r"Emolumentos",
    ],
    "ana_fee": [
        r"Taxa\s+A\.?N\.?A\.?",
    ],
    "operational_fee": [
        r"Taxa\s+Operacional",
    ],
    "execution_fee": [
        r"^Execu[cç][aã]o$",
    ],
    "custody_fee": [
        r"Taxa\s+de\s+Cust[oó]dia",
    ],
    "taxes": [
        r"^Impostos$",
        r"^ISS(?:\s*\([^)]*\))?$",
    ],
    "other_fee": [
        r"^Outras$",
        r"Outras\s+Bovespa",
    ],
    "irrf": [
        r"I\.?R\.?R\.?F",
    ],
}

# Rótulos usados para validação matemática; nunca entram no rateio.
TOTAL_LABEL_PATTERNS = {
    "net_operations":    r"Valor\s+l[ií]quido\s+das\s+opera[cç][õo]es",
    "total_cblc":        r"Total\s+CBLC",
    "total_bovespa":     r"Total\s+Bovespa\s*/\s*Soma|Total\s+Bolsa",
    "total_depositaria": r"Total\s+Deposit[aá]ria",
    "total_costs":       r"Total\s+Corretagem|Total\s+Custos",
    "liquido_para":      r"L[ií]quido\s+para",
}


###############################################################################
# FUNÇÕES AUXILIARES GERAIS
###############################################################################

class PdfPasswordRequiredError(RuntimeError):
    """Erro específico para PDF protegido por senha ou senha inválida."""
    pass


class BrokerageNoteValidationError(RuntimeError):
    """Erro quando a nota foi lida, mas não pertence à apuração informada."""
    pass


def resource_path(relative_path: str) -> str:
    """Retorna o caminho absoluto do recurso (funciona em .py e .exe/PyInstaller)."""
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_path, relative_path)


def get_data_path(filename: str) -> str:
    """Caminho para arquivos persistentes (json, config, etc.)."""
    if getattr(sys, "frozen", False):
        base_dir = os.path.dirname(sys.executable)
    else:
        base_dir = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_dir, filename)


def load_json_dict(path: str) -> Dict[str, str]:
    """Carrega um dicionário JSON do disco; retorna {} em caso de falha."""
    if not os.path.exists(path):
        return {}
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        if isinstance(data, dict):
            return {str(k): str(v) for k, v in data.items()}
        return {}
    except Exception:
        return {}


def save_json_dict(path: str, data: Dict[str, str]) -> None:
    """Salva um dicionário como JSON no disco com troca atômica."""
    dir_path = os.path.dirname(path)
    if dir_path:
        os.makedirs(dir_path, exist_ok=True)
    tmp_path = path + ".tmp"
    with open(tmp_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2, sort_keys=True)
    os.replace(tmp_path, path)


def _extract_cpf_from_pdf(pdf_path: str, password: Optional[str] = None) -> Optional[str]:
    """Extrai o primeiro CPF encontrado no texto do PDF (via PyMuPDF/fitz).

    Retorna somente os 11 dígitos do CPF, ou None se não encontrar.
    """
    cpf_regex = r"\b\d{3}\.?\d{3}\.?\d{3}-?\d{2}\b"
    try:
        doc = fitz.open(pdf_path)
        if password:
            try:
                if hasattr(doc, "authenticate"):
                    doc.authenticate(password)
            except Exception:
                return None
        for page in doc:
            text = page.get_text("text")
            match = re.search(cpf_regex, text)
            if match:
                cpf = re.sub(r"\D", "", match.group(0))
                if len(cpf) == 11:
                    return cpf
        return None
    except Exception:
        return None


###############################################################################
# DETECÇÃO DE TIPO DE ATIVO GERAL
###############################################################################

def detect_asset_type(name: str, ticker: Optional[str]) -> Optional[str]:
    """Detecta o tipo do ativo a partir da descrição do título e do ticker.

    A detecção é feita por prioridade, do tipo mais específico ao mais genérico:

    Prioridade de detecção:
    1. ETF           (lista fixa de 112 tickers)
    2. FII / FIAGRO  (ticker termina em 11 + palavra-chave no nome)
    3. BDR / DRN     (ticker termina em 33/34, ou "BDR"/"DRN" no nome)
    4. UNITS         (lista fixa de 13 tickers)
    5. ON            (palavra "ON" no nome)
    6. PN            (palavra "PN" no nome)

    Retorna o identificador canônico (ON, PN, FII, FIAGRO, BDR, UNITS, ETF)
    ou None se não conseguir identificar.
    """
    name_upper = (name or "").upper().strip()
    ticker_upper = (ticker or "").upper().strip()

    # ── 1. ETF CONHECIDO POR TICKER ──────────────────────────────────────────
    # Verifica a lista de ETFs conhecidos.
    if ticker_upper in ETF_TICKERS_CONHECIDOS:
        return ASSET_TYPE_ETF

    # ── 2. FII / FIAGRO ──────────────────────────────────────────────────────
    # Ticker termina com "11" E nome contém palavra-chave de fundo (dupla condição).
    if ticker_upper.endswith("11"):
        fii_keywords = ("FII", "FDO", "FUNDO", "FIAGRO")
        if any(kw in name_upper for kw in fii_keywords):
            # Distingue FIAGRO de FII pelo nome
            if "FIAGRO" in name_upper:
                return ASSET_TYPE_FIAGRO
            return ASSET_TYPE_FII

    # ── 3. BDR / DRN ─────────────────────────────────────────────────────────
    # Tickers terminados em 34 (Nível III) ou 33 (Nível I/II) e/ou contêm "BDR"/"DRN" no nome.
    if re.search(r"\d{2}(33|34)$", ticker_upper):
        return ASSET_TYPE_BDR
    if re.search(r"\bBDR\b|\bDRN\b", name_upper):
        return ASSET_TYPE_BDR

    # ── 4. UNITS ─────────────────────────────────────────────────────────────
    # Verifica a lista de UNITS conhecidos.
    if ticker_upper in UNITS_TICKERS_CONHECIDOS:
        return ASSET_TYPE_UNITS

    # ── 5. ON (Ordinária) ────────────────────────────────────────────────────
    # A palavra "ON" aparece na descrição do título nas notas.
    # Exemplos: "VALE ON NM", "KLABIN S/A ON", "RRRP3 ON NM"
    # \b para evitar falso positivo com palavras que contenham "ON".
    if re.search(r"\bON\b", name_upper):
        return ASSET_TYPE_ON

    # ── 6. PN (Preferencial) ─────────────────────────────────────────────────
    # A palavra "PN" aparece na descrição do título nas notas.
    # Exemplos: "KLABIN S/A PN N2", "TAESA PN N2"
    # \b para evitar falso positivo com palavras que contenham "PN".
    if re.search(r"\bPN\b", name_upper):
        return ASSET_TYPE_PN

    # Não foi possível identificar automaticamente → retorna None
    # O chamador deve solicitar ao usuário via modal.
    return None


def normalize_b3_ticker(ticker: Optional[str]) -> Optional[str]:
    """Normaliza ticker B3 removendo o F final de mercado fracionário.

    Ex.: BBAS3F e KBLN11F representam o mesmo ativo negociado à vista como
    BBAS3 e KBLN11. A remoção só acontece quando o ticker segue o padrão de
    ativo B3 com 4 letras, 1 ou 2 dígitos e F no final.
    """
    if ticker is None:
        return None
    ticker_upper = str(ticker).strip().upper()
    if not ticker_upper:
        return None
    if re.fullmatch(r"[A-Z]{4}\d{1,2}F", ticker_upper):
        return ticker_upper[:-1]
    return ticker_upper


# Função auxiliar do orquestrador: identifica corretora para escolher parser específico.
def _detect_broker(text: str) -> str:
    """Identifica a corretora pelo CNPJ ou nome presente no texto do PDF.

    Retorna uma string identificadora da corretora (ex.: "inter"),
    ou "unknown" se não reconhecer.
    """
    text_digits_only = re.sub(r"\D", "", text)
    for cnpj_fragment, broker_name in BROKER_CNPJ_MAP.items():
        if cnpj_fragment in text_digits_only:
            return broker_name
    text_upper = (text or "").upper()
    if "GENIAL" in text_upper:
        return "genial"
    if "INTER DTVM" in text_upper or "BANCOINTER" in text_upper:
        return "inter"
    return "unknown"

# Função auxiliar dos parsers: extrai texto bruto do PDF com PyMuPDF.
def _extract_text_from_pdf(pdf_path: str, password: Optional[str] = None) -> str:
    """Extrai texto bruto de todas as páginas do PDF usando PyMuPDF.

    Esse texto é usado por:
    - detecção de corretora;
    - parser genérico;
    - tela de revisão manual assistida.

    Observação:
    PDFs podem retornar texto fora da ordem visual, por isso o parser genérico
    deve ser tratado como fallback e validado antes de salvar a apuração.
    """
    try:
        doc = fitz.open(pdf_path)
        if password:
            try:
                doc.authenticate(password)
            except Exception:
                pass
        return "\n".join(page.get_text("text") for page in doc)
    except Exception as exc:
        raise RuntimeError(f"Não foi possível extrair texto do PDF: {exc}") from exc

# Função auxiliar dos parsers: extrai CPF do texto bruto da nota.
def _extract_cpf_from_text(text: str) -> Optional[str]:
    """Extrai o primeiro CPF encontrado em um texto bruto."""
    match = re.search(r"\b(\d{3}[\.\s]?\d{3}[\.\s]?\d{3}[\-\s]?\d{2})\b", text or "")
    if not match:
        return None
    cpf = re.sub(r"\D", "", match.group(1))
    return cpf if len(cpf) == 11 else None

# Função auxiliar dos parsers: extrai a data de pregão do texto bruto.
def _extract_ref_date_from_text(text: str) -> date:
    """Tenta localizar a data do pregão no texto bruto da nota.

    Aceita variações comuns:
    - Data pregão: DD/MM/AAAA
    - Data do pregão DD/MM/AAAA
    - cabeçalho com "Folha Data pregão" seguido da data
    """
    patterns = [
        r"Data\s+(?:do\s+)?preg[aã]o[:\s]+(\d{2}/\d{2}/\d{4})",
        r"Folha\s+Data\s+preg[aã]o\s+.*?(\d{2}/\d{2}/\d{4})",
        r"\b(\d{2}/\d{2}/\d{4})\b",
    ]

    for pattern in patterns:
        match = re.search(pattern, text or "", re.IGNORECASE | re.DOTALL)
        if match:
            return _parse_date_br(match.group(1))

    raise RuntimeError("Não foi possível localizar a data do pregão no texto da nota.")

# Função auxiliar do orquestrador: valida mês/ano e CPF das linhas extraídas.
def _validate_rows_against_month_cpf(
    rows: List[dict],
    filename: str,
    expected_year: int,
    expected_month: int,
    expected_cpf_digits: str,
    pdf_path: str,
    password: Optional[str] = None,
) -> None:
    """Valida data e CPF das linhas extraídas por parser alternativo/manual."""
    for r in rows:
        rd: date = r["ref_date"]
        if rd.year != expected_year or rd.month != expected_month:
            raise BrokerageNoteValidationError(
                f"A nota '{filename}' possui data de pregão {rd:%d/%m/%Y}, "
                f"que não pertence ao mês {expected_month:02d}/{expected_year}."
            )

        if expected_cpf_digits:
            note_cpf = r.get("cpf") or _extract_cpf_from_pdf(pdf_path, password=password)
            if note_cpf is not None and note_cpf != expected_cpf_digits:
                raise BrokerageNoteValidationError(
                    f"A nota '{filename}' possui CPF {formatar_cpf(note_cpf)}, "
                    f"diferente do CPF informado {formatar_cpf(expected_cpf_digits)}."
                )

# Função auxiliar do orquestrador: confirma se as linhas têm operações utilizáveis.
def _rows_have_valid_operations(rows: List[dict]) -> bool:
    """Retorna True se as linhas extraídas têm operação com quantidade, preço e total utilizáveis."""

    for row in rows:
        amount = row.get("amount")
        unit_price = row.get("unit_price")
        total_value = row.get("total_value")
        if amount is not None and amount > 0 and unit_price is not None and unit_price > 0:
            if total_value is None or total_value <= 0:
                total_value = amount * unit_price
            if total_value > 0:
                return True
    return False

# Função auxiliar da Camada 1: valida se o CorrePy trouxe transações utilizáveis.
def _correpy_notes_have_valid_transactions(notes: List) -> bool:
    """Retorna True se o CorrePy extraiu ao menos uma transação utilizável."""

    for note in notes:
        for tx in getattr(note, "transactions", []) or []:
            amount = getattr(tx, "amount", None)
            unit_price = getattr(tx, "unit_price", None)
            total_value = getattr(tx, "total_value", None)
            if amount is not None and amount > 0 and unit_price is not None and unit_price > 0:
                if total_value is None or total_value <= 0:
                    total_value = amount * unit_price
                if total_value > 0:
                    return True
    return False


def _apply_financial_summary_to_rows(rows: List[dict], financial: FinancialSummary) -> None:
    """Copia o Resumo Financeiro normalizado para as linhas brutas do parser."""
    for row in rows:
        row["settlement_fee_note"] = financial.settlement_fee
        row["emoluments_note"] = financial.emoluments
        row["registration_fee_note"] = financial.registration_fee
        row["transfer_fee_note"] = financial.transfer_fee
        row["ana_fee_note"] = financial.ana_fee
        row["operational_fee_note"] = financial.operational_fee
        row["execution_fee_note"] = financial.execution_fee
        row["custody_fee_note"] = financial.custody_fee
        row["taxes_note"] = financial.taxes
        row["other_fee_note"] = financial.other_fee
        row["irrf"] = financial.irrf


def _extract_and_apply_financial_summary(
    rows: List[dict],
    lines: List[str],
    broker: str,
    warning_prefix: Optional[str] = None,
) -> FinancialSummary:
    """Extrai o Resumo Financeiro da nota e aplica o resultado às linhas lidas."""
    total_value_note = sum(r["total_value"] for r in rows)
    financial = extract_financial_summary(lines, operation_total=total_value_note, broker=broker)
    _apply_financial_summary_to_rows(rows, financial)
    if warning_prefix and financial.warnings:
        print(f"{warning_prefix} {financial.warnings} — nota pode ter taxas incorretas")
    return financial


###############################################################################
# CAMADA 2 — PARSERS ESPECÍFICOS POR CORRETORA (Inter, Genial)
###############################################################################

# Parser específico Inter: mantém a leitura de operações por tabela pdfplumber.
def _parse_inter_pdfplumber(pdf_path: str, password: Optional[str] = None) -> List[dict]:
    """Parser específico para o layout de nota do Banco Inter (Inter DTVM Ltda.).

    Layout da tabela do Inter:
        Praça | C/V | Tipo Mercado | Especificação do Título | OBS(*) | Quantidade | Preço Liquidação (R$) | Compra/Venda (R$) | D/C

    Extrai também:
    - Data do pregão do cabeçalho
    - CPF do cliente
    - Taxas do Resumo Financeiro (liquidação, emolumentos, registro, transferência de ativos)

    Retorna lista de dicts com:
        ref_date, ticker, name, transaction_type, amount, unit_price,
        total_value, settlement_fee_note, emoluments_note,
        registration_fee_note, transfer_fee_note, ana_fee_note,
        other_fee_note, irrf, cpf
    """
    open_kwargs = {"password": password} if password else {}

    with pdfplumber.open(pdf_path, **open_kwargs) as pdf:
        full_text = "\n".join(page.extract_text() or "" for page in pdf.pages)
    lines = [
        re.sub(r"\s+", " ", raw_line).strip()
        for raw_line in full_text.splitlines()
        if raw_line and raw_line.strip()
    ]

    # Extrai data do pregão: "Data pregão: DD/MM/AAAA"
    date_match = re.search(r"Data\s+preg[aã]o[:\s]+(\d{2}/\d{2}/\d{4})", full_text, re.IGNORECASE)
    if not date_match:
        raise RuntimeError("Parser Inter: não foi possível localizar a data do pregão.")
    ref_date = _parse_date_br(date_match.group(1))

    # Extrai CPF do cliente
    cpf_match = re.search(r"\b(\d{3}[\.\s]?\d{3}[\.\s]?\d{3}[\-\s]?\d{2})\b", full_text)
    cpf_digits = re.sub(r"\D", "", cpf_match.group(1)) if cpf_match else None

    settlement_fee = Decimal("0")
    registration_fee = Decimal("0")
    transfer_fee = Decimal("0")
    ana_fee = Decimal("0")
    emoluments = Decimal("0")
    other_fee = Decimal("0")
    irrf_note = Decimal("0")

    # Extrai operações da tabela
    operations: List[dict] = []

    with pdfplumber.open(pdf_path, **open_kwargs) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables:
                if not table:
                    continue
                for row in table:
                    row_clean = [str(c).strip() if c is not None else "" for c in row]
                    if len(row_clean) < 7:
                        continue

                    # Coluna C/V identifica linha de operação
                    cv_col = row_clean[1].upper().strip()
                    if cv_col not in ("C", "V"):
                        continue

                    # "Especificação do Título": "RRRP3 ON NM"
                    especificacao = row_clean[3].strip()
                    if not especificacao:
                        continue

                    partes_spec = especificacao.split()
                    ticker = normalize_b3_ticker(partes_spec[0]) if partes_spec else None
                    name   = especificacao  # nome completo para detect_asset_type

                    amount     = _parse_integer_br(row_clean[5])
                    unit_price = _parse_decimal_br(row_clean[6])
                    if amount is None or amount <= 0:
                        continue
                    if unit_price is None or unit_price <= 0:
                        continue

                    operations.append({
                        "ref_date": ref_date,
                        "ticker": ticker,
                        "name": name,
                        "transaction_type": "buy" if cv_col == "C" else "sell",
                        "amount": amount,
                        "unit_price": unit_price,
                        "total_value": amount * unit_price,
                        "settlement_fee_note": settlement_fee,
                        "emoluments_note": emoluments,
                        "registration_fee_note": registration_fee,
                        "transfer_fee_note": transfer_fee,
                        "ana_fee_note": ana_fee,
                        "other_fee_note": other_fee,
                        "irrf": irrf_note,
                        "cpf": cpf_digits,
                    })

    if not operations:
        raise RuntimeError("Parser Inter (pdfplumber): nenhuma operação encontrada na tabela da nota.")

    _extract_and_apply_financial_summary(operations, lines, broker="inter")

    return operations


# Helper legado: extrai taxa próxima ao rótulo, usado pela taxa de transferência do CorrePy.
def _extract_fee_near_label(lines: List[str], label_pattern: str) -> Decimal:
    """Extrai taxa em layouts verticais onde o valor pode vir antes ou depois do rótulo."""
    for idx, line in enumerate(lines):
        if not re.search(label_pattern, line, re.IGNORECASE):
            continue

        candidates: List[Tuple[int, Decimal]] = []
        for distance in range(1, 3):
            for neighbor_idx in (idx - distance, idx + distance):
                if neighbor_idx < 0 or neighbor_idx >= len(lines):
                    continue
                value = _parse_decimal_br(lines[neighbor_idx])
                if value is None or value < 0 or value > Decimal("1000"):
                    continue
                candidates.append((distance, value))

        if not candidates:
            continue
        candidates.sort(key=lambda item: item[0])
        return candidates[0][1]

    return Decimal("0")


# Helper da Camada 1: extrai taxa manual por data para complementar notas CorrePy.
def _extract_fee_by_ref_date_from_pdf(
    pdf_path: str,
    password: Optional[str],
    label_pattern: str,
) -> Dict[date, Decimal]:
    """Extrai uma taxa por data de pregão em PDFs com várias notas no mesmo arquivo."""
    fees_by_date: Dict[date, Decimal] = {}

    doc = fitz.open(pdf_path)
    if password:
        try:
            doc.authenticate(password)
        except Exception:
            pass

    for page in doc:
        page_text = page.get_text("text")
        try:
            ref_date = _extract_ref_date_from_text(page_text)
        except RuntimeError:
            continue

        lines = [
            re.sub(r"\s+", " ", raw_line).strip()
            for raw_line in page_text.splitlines()
            if raw_line and raw_line.strip()
        ]
        fee = _extract_fee_from_label_line(page_text, label_pattern)
        if fee == 0:
            fee = _extract_fee_near_label(lines, label_pattern)

        current = fees_by_date.get(ref_date)
        if current is None or (current == 0 and fee != 0):
            fees_by_date[ref_date] = fee

    doc.close()
    return fees_by_date


# Parser específico Genial: mantém a leitura de operações em blocos verticais.
def _parse_genial_pdfplumber(pdf_path: str, password: Optional[str] = None) -> List[dict]:
    """Parser específico para notas Genial com operações em blocos verticais."""
    rows: List[dict] = []

    doc = fitz.open(pdf_path)
    if password:
        try:
            doc.authenticate(password)
        except Exception:
            pass

    for page in doc:
        page_text = page.get_text("text")
        if "GENIAL" not in page_text.upper():
            continue

        ref_date = _extract_ref_date_from_text(page_text)
        cpf_digits = _extract_cpf_from_text(page_text)
        lines = [
            re.sub(r"\s+", " ", raw_line).strip()
            for raw_line in page_text.splitlines()
            if raw_line and raw_line.strip()
        ]

        page_rows = _parse_vertical_operation_blocks(
            lines=lines,
            ref_date=ref_date,
            cpf_digits=cpf_digits,
            settlement_fee=Decimal("0"),
            registration_fee=Decimal("0"),
            transfer_fee=Decimal("0"),
            ana_fee=Decimal("0"),
            emoluments=Decimal("0"),
            other_fee=Decimal("0"),
            irrf_note=Decimal("0"),
            start_mode="inline_cv",
        )

        if page_rows:
            _extract_and_apply_financial_summary(page_rows, lines, broker="genial")
            rows.extend(page_rows)

    if not rows:
        raise RuntimeError("Parser Genial: nenhuma operação encontrada na nota.")

    return rows


def _extract_fee_from_label_line(text: str, label_pattern: str) -> Decimal:
    """Extrai o último valor monetário da linha que contém o rótulo da taxa."""
    money_pattern = r"\d{1,3}(?:\.\d{3})*,\d+|\d+,\d+"
    for line in text.splitlines():
        if not re.search(label_pattern, line, re.IGNORECASE):
            continue
        values = re.findall(money_pattern, line)
        if not values:
            continue
        val = _parse_decimal_br(values[-1])
        return val if val is not None else Decimal("0")
    return Decimal("0")


def _parse_decimal_br(value: str) -> Optional[Decimal]:
    """Converte string BR (1.000,50) ou EN (1000.50) para Decimal.

    Retorna None se a conversão falhar.
    """
    if not value:
        return None
    v = str(value).strip()
    v = re.sub(r"[^\d,\.\-]", "", v)
    if not v:
        return None
    if "," in v and "." in v:
        if v.rfind(",") > v.rfind("."):
            v = v.replace(".", "").replace(",", ".")
        else:
            v = v.replace(",", "")
    elif "," in v:
        v = v.replace(".", "").replace(",", ".")
    try:
        return Decimal(v)
    except InvalidOperation:
        return None


def _parse_integer_br(value: str) -> Optional[Decimal]:
    """Converte inteiro em formato BR com separador de milhar (ex.: 1.000)."""
    if not value:
        return None
    v = re.sub(r"\D", "", str(value))
    if not v:
        return None
    return Decimal(v)


def _money_values_from_line(line: str) -> List[Decimal]:
    """Extrai todos os valores monetários de uma linha em formato brasileiro."""
    money_pattern = r"-?\d{1,3}(?:\.\d{3})*,\d+|-?\d+,\d+"
    values: List[Decimal] = []
    for raw_value in re.findall(money_pattern, line or ""):
        value = _parse_decimal_br(raw_value)
        if value is not None:
            values.append(value)
    return values


def _last_money_from_line(line: str) -> Optional[Decimal]:
    """Retorna o último valor monetário da linha, ignorando sufixos D/C."""
    values = _money_values_from_line(line)
    return values[-1] if values else None


def _line_without_money_and_dc(line: str) -> str:
    """Remove valores, R$ e D/C para comparar apenas o texto do rótulo."""
    clean = re.sub(r"R\$", " ", line or "", flags=re.IGNORECASE)
    clean = re.sub(r"-?\d{1,3}(?:\.\d{3})*,\d+|-?\d+,\d+", " ", clean)
    clean = re.sub(r"\b[DC]\b", " ", clean, flags=re.IGNORECASE)
    clean = re.sub(r"\s+", " ", clean).strip()
    return clean


def _label_matches(line: str, patterns: List[str]) -> bool:
    """Confere se uma linha bate com algum padrão de rótulo financeiro."""
    label_text = _line_without_money_and_dc(line)
    return any(re.search(pattern, label_text, re.IGNORECASE) for pattern in patterns)


def _financial_label_key(line: str) -> Optional[str]:
    """Identifica o campo canônico de uma linha de rótulo financeiro."""
    for field_name, patterns in FEE_LABEL_PATTERNS.items():
        if _label_matches(line, patterns):
            return field_name
    for field_name, pattern in TOTAL_LABEL_PATTERNS.items():
        if _label_matches(line, [pattern]):
            return field_name
    return None


def _is_value_only_line(line: str) -> bool:
    """Retorna True quando a linha contém somente um valor monetário e D/C opcional."""
    return bool(re.fullmatch(
        r"\s*(?:R\$)?\s*-?(?:\d{1,3}(?:\.\d{3})*|\d+),\d+\s*(?:[DC])?\s*",
        line or "",
        flags=re.IGNORECASE,
    ))


def _set_summary_field(summary: FinancialSummary, field_name: str, value: Decimal) -> None:
    """Preenche um campo de FinancialSummary somando quando o rótulo se repetir."""
    current = getattr(summary, field_name, Decimal("0"))
    setattr(summary, field_name, current + value)


def _try_fees_same_line(lines: List[str]) -> FinancialSummary:
    """Estratégia A: extrai taxas quando valor e rótulo aparecem na mesma linha."""
    summary = FinancialSummary(source="inline")
    seen_fee_lines = set()

    for idx, line in enumerate(lines):
        for field_name, patterns in FEE_LABEL_PATTERNS.items():
            if not _label_matches(line, patterns):
                continue
            if field_name == "irrf" and re.search(r"base\s+R\$", line, re.IGNORECASE):
                projection = re.search(r"Projeç[aã]o\s+R\$\s*([\d.,-]+)", line, re.IGNORECASE)
                if projection:
                    value = _parse_decimal_br(projection.group(1))
                else:
                    values = _money_values_from_line(line)
                    if len(values) <= 1:
                        continue
                    else:
                        value = values[-1]
                if value is None or value < 0:
                    continue
                _set_summary_field(summary, field_name, value)
                continue
            value = _last_money_from_line(line)
            if value is None or value < 0:
                continue
            normalized_line = re.sub(r"\s+", " ", line).strip().lower()
            seen_key = (field_name, value, normalized_line)
            if seen_key in seen_fee_lines:
                continue
            seen_fee_lines.add(seen_key)
            _set_summary_field(summary, field_name, value)

    return summary


def _try_fees_positional(lines: List[str], operation_total: Optional[Decimal] = None) -> FinancialSummary:
    """Estratégia B: associa rótulos e valores por posição dentro do Resumo Financeiro."""
    for idx, line in enumerate(lines):
        if _label_matches(line, FEE_LABEL_PATTERNS["settlement_fee"]):
            if _next_valid_value_line(lines, idx) is not None:
                return FinancialSummary(source="positional", warnings=["next_line_layout_detected"])

    start_idx = None
    end_idx = None

    for idx, line in enumerate(lines):
        if start_idx is None and _label_matches(line, [TOTAL_LABEL_PATTERNS["net_operations"]]):
            start_idx = idx
            continue
        if start_idx is not None and _label_matches(line, [TOTAL_LABEL_PATTERNS["liquido_para"]]):
            end_idx = idx
            break

    if start_idx is None:
        return FinancialSummary(source="positional", warnings=["financial_block_not_found"])

    block = lines[start_idx:end_idx] if end_idx is not None else lines[start_idx:]
    labels: List[str] = []
    values: List[Decimal] = []

    for line in block:
        label_key = _financial_label_key(line)
        if label_key is not None:
            labels.append(label_key)
        if _is_value_only_line(line):
            value = _last_money_from_line(line)
            if value is not None:
                values.append(value)

    mapped: Dict[str, Decimal] = {}

    if operation_total is not None:
        zero_fee_candidate: Optional[FinancialSummary] = None
        for offset, value in enumerate(values):
            if abs(value - operation_total) > Decimal("0.05"):
                continue
            if offset + 2 >= len(values):
                continue
            settlement_candidate = values[offset + 1]
            registration_candidate = values[offset + 2]
            expected_cblc = value + settlement_candidate + registration_candidate
            total_cblc_found = any(
                abs(candidate - expected_cblc) <= Decimal("0.05")
                for candidate in values[offset + 3:]
            )
            if not total_cblc_found:
                continue

            summary = FinancialSummary(
                settlement_fee=settlement_candidate,
                registration_fee=registration_candidate,
                source="positional",
            )
            summary.irrf = _try_fees_next_line(lines).irrf
            if summary.has_any_fee():
                return summary
            if zero_fee_candidate is None:
                zero_fee_candidate = summary
        if zero_fee_candidate is not None:
            return zero_fee_candidate

    for label_key, value in zip(labels, values):
        if label_key in mapped and not label_key.startswith("total_"):
            mapped[label_key] += value
        else:
            mapped[label_key] = value

    net_operations = mapped.get("net_operations")
    total_cblc = mapped.get("total_cblc")
    settlement_fee = mapped.get("settlement_fee", Decimal("0"))
    registration_fee = mapped.get("registration_fee", Decimal("0"))

    if net_operations is None or total_cblc is None:
        return FinancialSummary(source="positional", warnings=["cblc_validation_failed"])

    expected_cblc = net_operations + settlement_fee + registration_fee
    if abs(expected_cblc - total_cblc) > Decimal("0.05"):
        return FinancialSummary(source="positional", warnings=["cblc_validation_failed"])

    summary = FinancialSummary(source="positional")
    for field_name in FEE_LABEL_PATTERNS:
        value = mapped.get(field_name)
        if value is not None and value >= 0:
            setattr(summary, field_name, value)

    if summary.irrf == 0:
        summary.irrf = _try_fees_next_line(lines).irrf

    return summary


def _next_valid_value_line(lines: List[str], start_idx: int) -> Optional[Decimal]:
    """
    Retorna o valor monetário da PRÓXIMA linha imediatamente válida após start_idx.
    Uma linha válida é: não vazia, não contém apenas 'D' ou 'C', e contém um número monetário.
    Olha APENAS A PRÓXIMA linha não-vazia — não olha 2 linhas à frente nem para trás.
    Se a próxima linha válida não contiver número monetário, retorna None.
    """
    for candidate in lines[start_idx + 1:]:
        candidate = (candidate or "").strip()
        if not candidate:
            continue
        if re.fullmatch(r"[DC]", candidate, flags=re.IGNORECASE):
            continue
        return _last_money_from_line(candidate)
    return None


def _try_fees_next_line(lines: List[str]) -> FinancialSummary:
    """Estratégia C: extrai taxas quando o valor está na próxima linha válida do rótulo."""
    summary = FinancialSummary(source="next_line")

    for idx, line in enumerate(lines):
        for field_name, patterns in FEE_LABEL_PATTERNS.items():
            if not _label_matches(line, patterns):
                continue
            value = _next_valid_value_line(lines, idx)
            if value is None or value < 0:
                continue
            _set_summary_field(summary, field_name, value)

    return summary


def extract_financial_summary(
    lines: List[str],
    operation_total: Optional[Decimal] = None,
    broker: Optional[str] = None,
) -> FinancialSummary:
    """Extrai o Resumo Financeiro SINACOR usando estratégias em cascata."""
    inline = _try_fees_same_line(lines)
    if inline.has_any_fee():
        if inline.irrf == 0:
            inline_irrf = _try_fees_next_line(lines).irrf
            if inline_irrf > 0:
                inline.irrf = inline_irrf
        return inline

    positional = _try_fees_positional(lines, operation_total=operation_total)
    if positional.has_any_fee() and "cblc_validation_failed" not in positional.warnings:
        return positional

    next_line = _try_fees_next_line(lines)
    if next_line.has_any_fee():
        return next_line

    return FinancialSummary(source="none", warnings=["no_fees_extracted"])


def _recover_amount_from_extracted_text(
    text: str,
    ticker: Optional[str],
    unit_price: Decimal,
) -> Optional[Decimal]:
    """Recupera quantidade quando o CorrePy retorna 0 para uma operação.

    Algumas notas extraem a tabela em linhas verticais:
        TGAR11
        1650
        84,70
        139.755,00
        C

    Nesses casos, procuramos o ticker e, nas linhas seguintes, uma quantidade
    inteira seguida pelo mesmo preço unitário informado pelo CorrePy.
    """
    ticker_upper = (ticker or "").upper().strip()
    if not ticker_upper:
        return None

    lines = [
        re.sub(r"\s+", " ", line).strip()
        for line in text.splitlines()
        if line and line.strip()
    ]

    for idx, line in enumerate(lines):
        if not re.search(rf"\b{re.escape(ticker_upper)}\b", line.upper()):
            continue

        window = lines[idx + 1:idx + 10]
        for pos in range(len(window) - 1):
            amount = _parse_integer_br(window[pos])
            price = _parse_decimal_br(window[pos + 1])
            if amount is None or amount <= 0 or price is None:
                continue
            if price == unit_price:
                return amount

    return None


def _parse_date_br(date_str: str) -> date:
    """Converte 'DD/MM/AAAA' em objeto date."""
    day, month, year = date_str.strip().split("/")
    return date(int(year), int(month), int(day))


def _build_operations_from_pdfplumber_rows(
    rows: List[dict],
    filename: str,
    broker: str,
) -> List[Operation]:
    """Converte as linhas brutas do parser pdfplumber em objetos Operation.

    O rateio de taxas é proporcional ao valor financeiro de cada operação,
    idêntico ao que é feito para notas SINACOR (CorrePy).
    O IRRF é imposto retido da nota: não entra nas taxas, mas é rateado
    proporcionalmente após o valor financeiro das operações estar definido.
    O campo asset_type é deixado None; será preenchido depois por resolve_asset_types().
    """
    if not rows:
        return []

    operations: List[Operation] = []
    groups: Dict[Tuple[date, Decimal, Decimal, Decimal, Decimal, Decimal, Decimal, Decimal, Decimal, Decimal, Decimal, Decimal], List[dict]] = {}
    for r in rows:
        key = (
            r["ref_date"],
            r.get("settlement_fee_note", Decimal("0")),
            r.get("emoluments_note", Decimal("0")),
            r.get("registration_fee_note", Decimal("0")),
            r.get("transfer_fee_note", Decimal("0")),
            r.get("ana_fee_note", Decimal("0")),
            r.get("operational_fee_note", Decimal("0")),
            r.get("execution_fee_note", Decimal("0")),
            r.get("custody_fee_note", Decimal("0")),
            r.get("taxes_note", Decimal("0")),
            r.get("other_fee_note", Decimal("0")),
            r.get("irrf", Decimal("0")),
        )
        groups.setdefault(key, []).append(r)

    for (
        _ref_date,
        settlement_fee,
        emoluments,
        registration_fee,
        transfer_fee,
        ana_fee,
        operational_fee,
        execution_fee,
        custody_fee,
        taxes,
        other_fee,
        irrf_note,
    ), group_rows in groups.items():
        total_value_note = sum(r["total_value"] for r in group_rows)
        if total_value_note == 0:
            raise RuntimeError(
                f"Parser pdfplumber ({broker}): valor total zerado no arquivo '{filename}'."
            )

        total_fees = (
            settlement_fee + registration_fee + ana_fee
            + emoluments + transfer_fee + operational_fee
            + execution_fee + custody_fee + taxes + other_fee
        )
        for r in group_rows:
            proportion = r["total_value"] / total_value_note
            operations.append(Operation(
                ref_date=r["ref_date"],
                ticker=normalize_b3_ticker(r["ticker"]),
                name=r["name"],
                transaction_type=r["transaction_type"],
                amount=r["amount"],
                unit_price=r["unit_price"],
                total_value=r["total_value"],
                allocated_fee=total_fees * proportion,
                irrf=irrf_note * proportion,
                note_file=filename,
                asset_type=None,
                parser_used=f"pdfplumber_{broker}",
            ))

    return operations

###############################################################################
# CAMADA 3 — PARSER GENÉRICO SINACOR (fallback para corretoras não mapeadas)
###############################################################################

# Parser genérico tabular: lê operações SINACOR quando a linha sai completa no texto.
def _parse_generic_text_bovespa(pdf_path: str, password: Optional[str] = None) -> List[dict]:
    """Parser genérico para notas Bovespa à vista quando CorrePy e parser específico falham.

    Estratégia:
    - extrai texto bruto com PyMuPDF;
    - procura linhas com C/V;
    - identifica quantidade, preço unitário, valor total e D/C pelo fim da linha;
    - tenta localizar ticker por padrão brasileiro (ex.: PETR4, HGLG11);
    - se não localizar ticker, deixa None para a sua janela `_pedir_ticker` resolver depois.

    Limitação intencional:
    - Este parser é voltado para operações à vista/fracionário de B3.
    - Futuros/BMF/WIN/WDO devem ter parser próprio, pois o cálculo tributário é diferente.
    """
    filename = os.path.basename(pdf_path)
    text = _extract_text_from_pdf(pdf_path, password=password)

    ref_date = _extract_ref_date_from_text(text)
    cpf_digits = _extract_cpf_from_text(text)

    lines = [
        re.sub(r"\s+", " ", raw_line).strip()
        for raw_line in text.splitlines()
        if raw_line and raw_line.strip()
    ]

    settlement_fee = Decimal("0")
    registration_fee = Decimal("0")
    transfer_fee = Decimal("0")
    ana_fee = Decimal("0")
    emoluments = Decimal("0")
    other_fee = Decimal("0")
    irrf_note = Decimal("0")

    money = r"(?:R\$\s*)?(?:\d{1,3}(?:\.\d{3})*|\d+),\d{2,6}"
    qty   = r"\d{1,12}(?:\.\d{3})*"

    # Captura o fim da linha:
    # ... QTD PRECO VALOR D/C
    end_pattern = re.compile(
        rf"(?P<before>.+?)\s+(?P<amount>{qty})\s+(?P<unit>{money})\s+(?P<total>{money})\s+(?P<dc>[DC])(?:\s|$)",
        re.IGNORECASE,
    )

    rows: List[dict] = []

    for raw_line in text.splitlines():
        line = re.sub(r"\s+", " ", raw_line).strip()
        if not line:
            continue

        line_upper = line.upper()

        # Ignora linhas de resumo/rodapé.
        if any(x in line_upper for x in ("SUBTOTAL", "RESUMO", "TOTAL", "LIQUIDO", "LÍQUIDO")):
            continue

        # Só tenta linhas com indicação operacional e mercado Bovespa/B3.
        if not re.search(r"\b[CV]\b", line_upper):
            continue
        if not re.search(r"BOVESPA|B3|VISTA|VIS|FRACION", line_upper):
            continue

        # Evita tentar interpretar futuros/BMF no parser genérico de ações.
        if re.search(r"\b(WIN|WDO|IND|DOL)\b|BM&F|BMF|FUTURO", line_upper):
            continue

        match = end_pattern.search(line_upper)
        if not match:
            continue

        before = match.group("before").strip()
        tokens = before.split()

        cv_idx = None
        for i, token in enumerate(tokens):
            if token in ("C", "V"):
                cv_idx = i
                break

        if cv_idx is None:
            continue

        cv = tokens[cv_idx]
        transaction_type = "buy" if cv == "C" else "sell"

        # Normalmente, após C/V vem o tipo de mercado (VISTA/VIS/FRACIONARIO).
        name_tokens = tokens[cv_idx + 2:] if len(tokens) > cv_idx + 2 else tokens[cv_idx + 1:]
        name = " ".join(name_tokens).strip()

        # Tenta localizar ticker brasileiro dentro da descrição.
        ticker_match = re.search(r"\b[A-Z]{4}\d{1,2}[A-Z]?\b", name)
        ticker = normalize_b3_ticker(ticker_match.group(0)) if ticker_match else None

        amount = _parse_integer_br(match.group("amount"))
        unit_price = _parse_decimal_br(match.group("unit"))
        total_value = _parse_decimal_br(match.group("total"))

        if amount is None or amount <= 0 or unit_price is None or unit_price <= 0:
            continue

        if total_value is None or total_value <= 0:
            total_value = amount * unit_price

        if not name:
            name = ticker or "ATIVO NAO IDENTIFICADO"

        rows.append({
            "ref_date": ref_date,
            "ticker": ticker,
            "name": name,
            "transaction_type": transaction_type,
            "amount": amount,
            "unit_price": unit_price,
            "total_value": total_value,
            "settlement_fee_note": settlement_fee,
            "emoluments_note": emoluments,
            "registration_fee_note": registration_fee,
            "transfer_fee_note": transfer_fee,
            "ana_fee_note": ana_fee,
            "other_fee_note": other_fee,
            "irrf": irrf_note,
            "cpf": cpf_digits,
        })

    if not rows:
        raise RuntimeError(f"Parser genérico tabular: nenhuma operação Bovespa encontrada em '{filename}'.")

    broker = _detect_broker(text)
    _extract_and_apply_financial_summary(
        rows,
        lines,
        broker=broker,
        warning_prefix="[AVISO parser genérico]",
    )

    return rows


# Parser genérico vertical: lê operações quando as colunas do PDF saem empilhadas.
def _parse_generic_vertical_bovespa(pdf_path: str, password: Optional[str] = None) -> List[dict]:
    """Parser genérico vertical para notas em que as colunas saem como linhas separadas."""

    filename = os.path.basename(pdf_path)
    text = _extract_text_from_pdf(pdf_path, password=password)
    ref_date = _extract_ref_date_from_text(text)
    cpf_digits = _extract_cpf_from_text(text)

    lines = [
        re.sub(r"\s+", " ", raw_line).strip()
        for raw_line in text.splitlines()
        if raw_line and raw_line.strip()
    ]

    settlement_fee = Decimal("0")
    registration_fee = Decimal("0")
    transfer_fee = Decimal("0")
    ana_fee = Decimal("0")
    emoluments = Decimal("0")
    other_fee = Decimal("0")
    irrf_note = Decimal("0")

    rows: List[dict] = []
    for start_mode in ("separate_cv", "leading_cv"):
        rows.extend(_parse_vertical_operation_blocks(
            lines=lines,
            ref_date=ref_date,
            cpf_digits=cpf_digits,
            settlement_fee=settlement_fee,
            registration_fee=registration_fee,
            transfer_fee=transfer_fee,
            ana_fee=ana_fee,
            emoluments=emoluments,
            other_fee=other_fee,
            irrf_note=irrf_note,
            start_mode=start_mode,
        ))

    if not rows:
        raise RuntimeError(f"Parser genérico vertical: nenhuma operação Bovespa encontrada em '{filename}'.")

    broker = _detect_broker(text)
    _extract_and_apply_financial_summary(
        rows,
        lines,
        broker=broker,
        warning_prefix="[AVISO parser genérico]",
    )

    return rows


# Função comum dos parsers verticais: converte blocos de texto em linhas de operação.
def _parse_vertical_operation_blocks(
    lines: List[str],
    ref_date: date,
    cpf_digits: Optional[str],
    settlement_fee: Decimal,
    registration_fee: Decimal,
    transfer_fee: Decimal,
    ana_fee: Decimal,
    emoluments: Decimal,
    other_fee: Decimal,
    irrf_note: Decimal,
    start_mode: str,
) -> List[dict]:
    """Lê operações em blocos verticais extraídos do PDF.

    start_mode:
    - "separate_cv": início em BOVESPA/B3, C/V na linha seguinte.
    - "inline_cv"  : início com C/V na mesma linha, ex.: B3 RV LISTADO V.
    - "leading_cv" : início com C/V em uma linha e B3 na linha seguinte.
    """
    money = r"(?:R\$\s*)?(?:\d{1,3}(?:\.\d{3})*|\d+),\d{2,6}"
    money_with_optional_dc = rf"{money}(?:\s*[DC])?"
    qty = r"\d{1,12}(?:\.\d{3})*"
    rows: List[dict] = []

    i = 0
    while i < len(lines):
        line_upper = lines[i].upper()

        if start_mode == "inline_cv":
            start_match = re.search(r"\bB3\s+RV\s+LISTADO\s+([CV])\b", line_upper)
            if not start_match:
                i += 1
                continue
            cv = start_match.group(1)
            market_idx = i + 1
            name_idx = i + 2
        elif start_mode == "leading_cv":
            cv = lines[i].strip().upper()
            if cv not in ("C", "V"):
                i += 1
                continue

            b3_idx = None
            for candidate_idx in range(i + 1, min(i + 6, len(lines))):
                if re.search(r"BOVESPA|B3", lines[candidate_idx], re.IGNORECASE):
                    b3_idx = candidate_idx
                    break
            if b3_idx is None or b3_idx + 2 >= len(lines):
                i += 1
                continue

            market_idx = b3_idx + 1
            name_idx = b3_idx + 2
        else:
            if not re.search(r"BOVESPA|B3", line_upper):
                i += 1
                continue
            if i + 2 >= len(lines):
                break
            cv = lines[i + 1].strip().upper()
            if cv not in ("C", "V"):
                i += 1
                continue
            market_idx = i + 2
            name_idx = i + 3

        if market_idx >= len(lines) or not re.search(r"VISTA|VIS|FRACION", lines[market_idx], re.IGNORECASE):
            i += 1
            continue

        name_parts: List[str] = []
        j = name_idx
        while j < len(lines):
            candidate = lines[j].strip()
            if candidate in ("@", "@#"):
                j += 1
                continue
            if re.fullmatch(qty, candidate):
                break
            if re.fullmatch(r"D[#\d]*", candidate.upper()):
                j += 1
                break
            if re.search(r"RESUMO|TOTAL|LIQUIDO|LÍQUIDO", candidate, re.IGNORECASE):
                break
            name_parts.append(candidate)
            j += 1

        if (
            not name_parts
            or j + 2 >= len(lines)
            or not re.fullmatch(qty, lines[j].strip())
            or not re.fullmatch(money, lines[j + 1].strip())
            or not re.fullmatch(money_with_optional_dc, lines[j + 2].strip(), re.IGNORECASE)
        ):
            i += 1
            continue

        total_line = lines[j + 2].strip()
        dc_idx = None
        if re.search(r"\b[DC]$", total_line, re.IGNORECASE):
            dc_idx = j + 2
        else:
            for candidate_idx in range(j + 3, min(j + 7, len(lines))):
                if lines[candidate_idx].strip().upper() in ("C", "D"):
                    dc_idx = candidate_idx
                    break
        if dc_idx is None:
            i += 1
            continue

        amount = _parse_integer_br(lines[j])
        unit_price = _parse_decimal_br(lines[j + 1])
        total_value = _parse_decimal_br(lines[j + 2])
        if amount is None or amount <= 0 or unit_price is None or unit_price <= 0:
            i = j + 4
            continue
        if total_value is None or total_value <= 0:
            total_value = amount * unit_price

        name = " ".join(name_parts).strip()
        ticker_match = re.search(r"\b[A-Z]{4}\d{1,2}[A-Z]?\b", name.upper())
        ticker = normalize_b3_ticker(ticker_match.group(0)) if ticker_match else None

        rows.append({
            "ref_date": ref_date,
            "ticker": ticker,
            "name": name or ticker or "ATIVO NAO IDENTIFICADO",
            "transaction_type": "buy" if cv == "C" else "sell",
            "amount": amount,
            "unit_price": unit_price,
            "total_value": total_value,
            "settlement_fee_note": settlement_fee,
            "emoluments_note": emoluments,
            "registration_fee_note": registration_fee,
            "transfer_fee_note": transfer_fee,
            "ana_fee_note": ana_fee,
            "other_fee_note": other_fee,
            "irrf": irrf_note,
            "cpf": cpf_digits,
        })
        i = dc_idx + 1

    return rows


###############################################################################
# CAMADA 4 — REVISÃO MANUAL ASSISTIDA (quando todos os parsers falham)
###############################################################################

def _parse_manual_review_lines(
    grid_rows: List[Dict[str, str]],
    taxas: Dict[str, Decimal],
    cpf_digits: Optional[str] = None,
) -> List[dict]:
    """Converte as linhas da grade visual em linhas de operação.

    Cada elemento de grid_rows é um dict com as chaves:
        data, c/v, ticker, nome, quantidade, preco, total

    Esses valores vêm dos campos Entry da tabela visual de _revisao_manual_nota().
    O campo 'total' é opcional: se vazio ou zero, é calculado como quantidade x preço.
    """
    rows: List[dict] = []

    for line_no, cell in enumerate(grid_rows, start=1):
        # Ignora linhas completamente em branco
        valores = [v.strip() for v in cell.values()]
        if not any(v for v in valores[:6]):   # os 6 primeiros campos obrigatórios
            continue

        # ── Data ──────────────────────────────────────────────────────────────
        data_str = cell.get("data", "").strip()
        if not data_str:
            raise RuntimeError(f"Linha {line_no}: campo 'Data' está vazio.")
        try:
            ref_date = _parse_date_br(data_str)
        except Exception:
            raise RuntimeError(
                f"Linha {line_no}: data '{data_str}' inválida. Use DD/MM/AAAA."
            )

        # ── C/V ───────────────────────────────────────────────────────────────
        cv = cell.get("cv", "").strip().upper()
        if cv not in ("C", "V"):
            raise RuntimeError(
                f"Linha {line_no}: C/V deve ser exatamente 'C' (compra) ou 'V' (venda). "
                f"Recebido: '{cv}'"
            )

        # ── Ticker ────────────────────────────────────────────────────────────
        ticker = normalize_b3_ticker(cell.get("ticker", "")) or ""
        if not ticker or not re.fullmatch(r"[A-Z0-9]{4,12}", ticker):
            raise RuntimeError(
                f"Linha {line_no}: ticker '{ticker}' inválido. "
                "Use 4 a 12 letras/números, ex.: PETR4, HGLG11."
            )

        # ── Nome ─────────────────────────────────────────────────────────────
        name = cell.get("nome", "").strip().upper() or ticker

        # ── Quantidade ───────────────────────────────────────────────────────
        amount = _parse_integer_br(cell.get("quantidade", ""))
        if amount is None or amount <= 0:
            raise RuntimeError(
                f"Linha {line_no}: quantidade '{cell.get('quantidade')}' inválida. "
                "Use um número inteiro positivo, ex.: 100"
            )

        # ── Preço ─────────────────────────────────────────────────────────────
        unit_price = _parse_decimal_br(cell.get("preco", ""))
        if unit_price is None or unit_price <= 0:
            raise RuntimeError(
                f"Linha {line_no}: preço '{cell.get('preco')}' inválido. "
                "Use vírgula como decimal, ex.: 28,50"
            )

        # ── Total (opcional) ──────────────────────────────────────────────────
        total_str = cell.get("total", "").strip()
        total_value = _parse_decimal_br(total_str) if total_str else None
        if total_value is None or total_value <= 0:
            total_value = amount * unit_price

        rows.append({
            "ref_date": ref_date,
            "ticker": ticker,
            "name": name,
            "transaction_type": "buy" if cv == "C" else "sell",
            "amount": amount,
            "unit_price": unit_price,
            "total_value": total_value,
            "settlement_fee_note":  taxas.get("settlement_fee",  Decimal("0")),
            "emoluments_note":      taxas.get("emoluments",      Decimal("0")),
            "registration_fee_note": taxas.get("registration_fee", Decimal("0")),
            "transfer_fee_note":    taxas.get("transfer_fee",    Decimal("0")),
            "ana_fee_note":         taxas.get("ana_fee",         Decimal("0")),
            "other_fee_note":       taxas.get("other_fee",       Decimal("0")),
            "irrf":                 taxas.get("irrf",            Decimal("0")),
            "cpf": cpf_digits,
        })

    if not rows:
        raise RuntimeError(
            "Nenhuma operação foi preenchida. "
            "Preencha ao menos uma linha na tabela antes de confirmar."
        )

    return rows


###############################################################################
# ORQUESTRADOR (cascata de tentativas de parser)
###############################################################################

# Orquestrador principal: tenta CorrePy, parser específico, genérico e revisão manual.
def read_brokerage_notes(
    pdf_path: str,
    expected_year: int,
    expected_month: int,
    expected_cpf_digits: str,
    password: Optional[str] = None,
    manual_review_cb: Optional[Callable[[str, str, List[str]], Optional[List[dict]]]] = None,
) -> Tuple[List, str]:
    """Lê e valida uma nota usando a cascata de parsers.

    Ordem aplicada:
        1) CorrePy padrão;
        2) parser específico da corretora detectada;
        3) parser genérico;
        4) tela de revisão manual assistida.

    Retorna:
        (notes_or_ops, source)
        - notes_or_ops: lista de BrokerageNote (CorrePy) OU list[Operation] dos parsers alternativos;
        - source: "correpy", "pdfplumber_<corretora>", "generic_text" ou "manual_review".
    """

    filename = os.path.basename(pdf_path)
    errors: List[str] = []

    # ── 1) CAMADA CORREPY ─────────────────────────────────────────────────────
    correpy_validating = False
    try:
        with open(pdf_path, "rb") as f:
            content = io.BytesIO(f.read())
            content.seek(0)

        parser = ParserFactory(brokerage_note=content, password=password).parse()
        notes = list(parser)

        if not notes:
            errors.append("CorrePy: retornou lista vazia.")
        else:
            correpy_validating = True
            all_dates_valid = True
            for note in notes:
                if not hasattr(note, "reference_date") or note.reference_date is None:
                    all_dates_valid = False
                    continue

                ref_date = note.reference_date
                if ref_date.year != expected_year or ref_date.month != expected_month:
                    raise BrokerageNoteValidationError(
                        f"A nota '{filename}' possui data de pregão {ref_date:%d/%m/%Y}, "
                        f"que não pertence ao mês {expected_month:02d}/{expected_year}."
                    )

                if expected_cpf_digits:
                    note_cpf = _extract_cpf_from_pdf(pdf_path, password=password)
                    if note_cpf is None:
                        raise RuntimeError(
                            f"Não foi possível identificar o CPF do cliente na nota '{filename}'."
                        )
                    if note_cpf != expected_cpf_digits:
                        raise BrokerageNoteValidationError(
                            f"A nota '{filename}' possui CPF {formatar_cpf(note_cpf)}, "
                            f"diferente do CPF informado {formatar_cpf(expected_cpf_digits)}."
                        )

            if not all_dates_valid:
                errors.append("CorrePy: uma ou mais notas foram montadas sem data de pregão válida.")
            elif not _correpy_notes_have_valid_transactions(notes):
                errors.append("CorrePy: não extraiu nenhuma transação com quantidade e preço utilizáveis.")
            else:
                return notes, "correpy"

    except InvalidPasswordException as e:
        raise PdfPasswordRequiredError(
            f"PDF protegido por senha ou senha inválida: {filename}"
        ) from e
    except PdfPasswordRequiredError:
        raise
    except BrokerageNoteValidationError:
        raise
    except RuntimeError as e:
        if correpy_validating:
            # RuntimeError aqui vem de validação de CPF/data após CorrePy montar a nota.
            # Nesse caso NÃO é seguro cair em parser alternativo.
            raise
        errors.append(f"CorrePy: {e}")
    except Exception as e:
        errors.append(f"CorrePy: {e}")

    # A partir daqui, precisa do texto bruto para detectar corretora, usar parser genérico
    # e preencher a tela manual se necessário.
    try:
        full_text = _extract_text_from_pdf(pdf_path, password=password)
    except Exception as e:
        full_text = ""
        errors.append(f"Extração de texto: {e}")

    broker = _detect_broker(full_text)

    # ── 2) PARSER ESPECÍFICO DA CORRETORA ─────────────────────────────────────
    if broker:
        parser_map = {
            "inter": _parse_inter_pdfplumber,
            "genial": _parse_genial_pdfplumber,
        }

        specific_parser = parser_map.get(broker)

        if specific_parser is None:
            errors.append(f"Parser específico: corretora '{broker}' sem parser cadastrado.")
        else:
            try:
                rows = specific_parser(pdf_path, password=password)
                if not _rows_have_valid_operations(rows):
                    raise RuntimeError("não extraiu nenhuma operação com quantidade, preço e valor utilizáveis.")

                _validate_rows_against_month_cpf(
                    rows=rows,
                    filename=filename,
                    expected_year=expected_year,
                    expected_month=expected_month,
                    expected_cpf_digits=expected_cpf_digits,
                    pdf_path=pdf_path,
                    password=password,
                )

                operations = _build_operations_from_pdfplumber_rows(rows, filename, broker)
                return operations, f"pdfplumber_{broker}"

            except BrokerageNoteValidationError:
                raise
            except Exception as e:
                errors.append(f"Parser específico ({broker}): {e}")

    # ── 3) PARSER GENÉRICO TABULAR ────────────────────────────────────────────
    try:
        rows = _parse_generic_text_bovespa(pdf_path, password=password)
        if not _rows_have_valid_operations(rows):
            raise RuntimeError("não extraiu nenhuma operação com quantidade, preço e valor utilizáveis.")

        _validate_rows_against_month_cpf(
            rows=rows,
            filename=filename,
            expected_year=expected_year,
            expected_month=expected_month,
            expected_cpf_digits=expected_cpf_digits,
            pdf_path=pdf_path,
            password=password,
        )

        operations = _build_operations_from_pdfplumber_rows(rows, filename, "generic")
        for op in operations:
            op.parser_used = "generic_text"
        return operations, "generic_text"

    except BrokerageNoteValidationError:
        raise
    except Exception as e:
        errors.append(f"Parser genérico tabular: {e}")

    # ── 4) PARSER GENÉRICO VERTICAL ───────────────────────────────────────────
    try:
        rows = _parse_generic_vertical_bovespa(pdf_path, password=password)
        if not _rows_have_valid_operations(rows):
            raise RuntimeError("não extraiu nenhuma operação com quantidade, preço e valor utilizáveis.")

        _validate_rows_against_month_cpf(
            rows=rows,
            filename=filename,
            expected_year=expected_year,
            expected_month=expected_month,
            expected_cpf_digits=expected_cpf_digits,
            pdf_path=pdf_path,
            password=password,
        )

        operations = _build_operations_from_pdfplumber_rows(rows, filename, "generic_vertical")
        for op in operations:
            op.parser_used = "generic_vertical"
        return operations, "generic_vertical"

    except BrokerageNoteValidationError:
        raise
    except Exception as e:
        errors.append(f"Parser genérico vertical: {e}")

    # ── 5) TELA DE REVISÃO MANUAL ASSISTIDA ───────────────────────────────────
    if manual_review_cb is not None:
        rows = manual_review_cb(filename, full_text, errors)
        if rows:
            if not _rows_have_valid_operations(rows):
                raise RuntimeError("Revisão manual: nenhuma operação com quantidade, preço e valor utilizáveis.")
            _validate_rows_against_month_cpf(
                rows=rows,
                filename=filename,
                expected_year=expected_year,
                expected_month=expected_month,
                expected_cpf_digits=expected_cpf_digits,
                pdf_path=pdf_path,
                password=password,
            )

            operations = _build_operations_from_pdfplumber_rows(rows, filename, "manual")
            for op in operations:
                op.parser_used = "manual_review"
            return operations, "manual_review"

        errors.append("Revisão manual: cancelada ou nenhuma operação informada.")

    raise RuntimeError(
        f"Não foi possível ler a nota '{filename}' por nenhuma camada.\n\n"
        + "\n".join(f"- {err}" for err in errors)
    )



###############################################################################
# Lógica de negócio: classificação, cálculo e carteira
###############################################################################

def is_fii(name: str, ticker: Optional[str]) -> bool:
    """Retorna True se o ativo for FII ou FIAGRO (ambas as condições devem ser atendidas).

    Regras:
    - Ticker termina com '11', E
    - Nome contém 'FII', 'FDO', 'FUNDO' ou 'FIAGRO'.

    Isso naturalmente exclui ETFs como BOVA11 (sem 'FII' no nome)
    e UNITS como SAPR11 (sem 'FII' no nome).
    """
    if not ticker:
        return False
    if ticker.endswith("11"):
        name_upper = name.upper()
        if any(keyword in name_upper for keyword in ("FII", "FDO", "FUNDO", "FIAGRO")):
            return True
    return False


def parse_month_folder(
    month_folder: str,
    expected_year: int,
    expected_month: int,
    expected_cpf_digits: str,
    password_map: Optional[Dict[str, Optional[str]]] = None,
    ask_password_cb: Optional[Callable[[str], Optional[str]]] = None,
    ask_quantity_cb: Optional[Callable[[str, date, str, str, Decimal], Optional[Decimal]]] = None,
    ask_manual_review_cb: Optional[Callable[[str, str, List[str]], Optional[List[dict]]]] = None,
) -> List[Operation]:
    """Varre a pasta do mês e extrai todas as operações.

    - Arquivos processados em ORDEM ALFABÉTICA para consistência entre execuções.
    - Senha solicitada individualmente por arquivo, com retry.
    - O campo asset_type de cada Operation fica None aqui;
      será preenchido depois por resolve_asset_types() na interface.
    """

    if password_map is None:
        password_map = {}

    operations: List[Operation] = []

    # sorted() garante ordem determinística independente do OS
    pdf_files = sorted(
        [f for f in os.listdir(month_folder) if f.lower().endswith(".pdf")]
    )

    for filename in pdf_files:
        pdf_path = os.path.join(month_folder, filename)
        senha_arquivo = password_map.get(filename)

        while True:
            try:
                result, source = read_brokerage_notes(
                    pdf_path,
                    expected_year=expected_year,
                    expected_month=expected_month,
                    expected_cpf_digits=expected_cpf_digits,
                    password=senha_arquivo,
                    manual_review_cb=ask_manual_review_cb,
                )
                password_map[filename] = senha_arquivo
                break
            except PdfPasswordRequiredError:
                if ask_password_cb is None:
                    raise RuntimeError(
                        f"O arquivo '{filename}' requer senha, mas não há mecanismo para solicitá-la."
                    )
                nova_senha = ask_password_cb(filename)
                if nova_senha is None:
                    raise RuntimeError(
                        f"Apuração cancelada: senha não informada para o arquivo '{filename}'."
                    )
                senha_arquivo = nova_senha

        if source == "correpy":
            notes = result
            raw_text_for_quantity: Optional[str] = None
            transfer_fee_pattern = r"Taxa\s+de\s+(?:Transfer[eê]ncia|Tranfer[eê]ncia|Transf\.?)\s+de\s+Ativos"
            transfer_fee_by_date = _extract_fee_by_ref_date_from_pdf(
                pdf_path,
                senha_arquivo,
                transfer_fee_pattern,
            )
            for note in notes:
                corrected_amounts: List[Decimal] = []
                corrected_tx_values: List[Decimal] = []
                total_value_note = Decimal(0)

                for tx in note.transactions:
                    amount = tx.amount
                    security_label = tx.security.ticker or tx.security.name or "ATIVO NAO IDENTIFICADO"
                    if amount is None or amount <= 0:
                        if raw_text_for_quantity is None:
                            raw_text_for_quantity = _extract_text_from_pdf(pdf_path, password=senha_arquivo)

                        recovered_amount = _recover_amount_from_extracted_text(
                            text=raw_text_for_quantity,
                            ticker=tx.security.ticker,
                            unit_price=tx.unit_price,
                        )
                        if recovered_amount is not None:
                            amount = recovered_amount
                        else:
                            if ask_quantity_cb is None:
                                raise RuntimeError(
                                    "Quantidade zerada encontrada, mas não há callback para solicitar ao usuário."
                                )
                            amount_user = ask_quantity_cb(
                                filename,
                                note.reference_date,
                                security_label,
                                tx.transaction_type.value,
                                tx.unit_price,
                            )
                            if amount_user is None:
                                raise RuntimeError(
                                    f"Operação cancelada: quantidade não informada para "
                                    f"{security_label} em {note.reference_date:%d/%m/%Y}."
                                )
                            amount = amount_user

                    if amount is None or amount <= 0:
                        raise RuntimeError(
                            f"Quantidade inválida para {security_label} "
                            f"em {note.reference_date:%d/%m/%Y} no arquivo '{filename}'."
                        )

                    corrected_amounts.append(amount)
                    tx_value = tx.unit_price * amount
                    corrected_tx_values.append(tx_value)
                    total_value_note += tx_value

                if total_value_note == 0:
                    raise RuntimeError(
                        f"O arquivo '{filename}' gerou valor total inválido (0)."
                    )

                if raw_text_for_quantity is None:
                    raw_text_for_quantity = _extract_text_from_pdf(pdf_path, password=senha_arquivo)
                transfer_fee_note = transfer_fee_by_date.get(note.reference_date, Decimal("0"))
                if note.reference_date not in transfer_fee_by_date:
                    transfer_fee_note = _extract_fee_from_label_line(
                        raw_text_for_quantity,
                        transfer_fee_pattern,
                    )
                if transfer_fee_note == 0 and note.reference_date not in transfer_fee_by_date:
                    transfer_fee_note = _extract_fee_near_label(
                        [
                            re.sub(r"\s+", " ", line).strip()
                            for line in raw_text_for_quantity.splitlines()
                            if line and line.strip()
                        ],
                        transfer_fee_pattern,
                    )

                total_fees = (
                    note.settlement_fee + note.registration_fee + note.ana_fee
                    + note.emoluments + transfer_fee_note + note.operational_fee + note.others
                )
                note_irrf = getattr(note, "irrf", Decimal("0")) or Decimal("0")

                for idx, tx in enumerate(note.transactions):
                    amount = corrected_amounts[idx]
                    tx_value = corrected_tx_values[idx]
                    proportion = (tx_value / total_value_note) if total_value_note else Decimal("0")

                    op = Operation(
                        ref_date=note.reference_date,
                        ticker=normalize_b3_ticker(tx.security.ticker),
                        name=tx.security.name,
                        transaction_type=tx.transaction_type.value,
                        amount=amount,
                        unit_price=tx.unit_price,
                        total_value=tx_value,
                        allocated_fee=total_fees * proportion,
                        irrf=note_irrf * proportion,
                        note_file=filename,
                        asset_type=None,  # preenchido depois
                        parser_used="correpy",
                    )
                    operations.append(op)
        else:
            operations.extend(result)

    if not operations:
        raise RuntimeError(
            "Nenhuma operação foi encontrada na pasta informada. "
            "Verifique se os PDFs estão no padrão SINACOR (ou Inter) e se pertencem ao mês selecionado."
        )

    return operations


def classify_operations(ops: List[Operation]) -> None:
    """Classifica cada operação como 'day', 'fii' ou 'swing'.

    Agrupa por (note_file, ref_date, ticker). Compra e venda da MESMA ação
    em notas DIFERENTES NÃO é detectado como day trade (comportamento intencional).

    Quando há day trade parcial, divide a operação em duas:
    parte "day" e parte "swing"/"fii".
    """
    new_ops: List[Operation] = []

    group_map: Dict[Tuple[str, date, str], List[Operation]] = {}
    for op in ops:
        op.ticker = normalize_b3_ticker(op.ticker)
        key = (op.note_file, op.ref_date, op.ticker)
        group_map.setdefault(key, []).append(op)

    for _, group_ops in group_map.items():
        buys  = [o for o in group_ops if o.transaction_type == "buy"]
        sells = [o for o in group_ops if o.transaction_type == "sell"]

        qty_day = min(sum(o.amount for o in buys), sum(o.amount for o in sells))

        if qty_day > 0:
            for ops_side in (buys, sells):
                remaining = qty_day
                for o in ops_side:
                    if remaining <= 0:
                        break
                    q = min(o.amount, remaining)
                    if q <= 0:
                        continue
                    ratio = q / o.amount if o.amount else Decimal(0)
                    day_op = deepcopy(o)
                    day_op.amount       = q
                    day_op.total_value  = q * o.unit_price
                    day_op.allocated_fee = o.allocated_fee * ratio
                    day_op.irrf         = (o.irrf or Decimal(0)) * ratio
                    day_op.category     = "day"
                    new_ops.append(day_op)
                    o.amount       -= q
                    o.total_value   = o.amount * o.unit_price
                    o.allocated_fee -= day_op.allocated_fee
                    o.irrf          = (o.irrf or Decimal(0)) - day_op.irrf
                    remaining -= q

        for o in group_ops:
            if o.amount <= 0:
                continue
            o.category = (
                "fii"
                if o.asset_type in (ASSET_TYPE_FII, ASSET_TYPE_FIAGRO) or is_fii(o.name, o.ticker)
                else "swing"
            )
            new_ops.append(o)

    ops.clear()
    ops.extend(new_ops)


def process_results(
    ops: List[Operation],
    positions: Dict[str, List[Tuple[Decimal, Decimal]]],
    irrf_day_from_notes: Decimal = Decimal(0),
) -> Tuple[
    Decimal, Decimal, Decimal, Decimal,  # result_swing_isento, result_swing_tributavel, result_day, result_fii
    Decimal, Decimal, Decimal,           # irrf_swing, irrf_day, irrf_fii
    Decimal, Decimal,                    # vendas_on_pn, vendas_outros (para G20 e regra dos 20k)
    Dict[str, List[Tuple[Decimal, Decimal]]],
    bool,
]:
    """Calcula os resultados líquidos separados por tipo de ativo.

    O custo das posições é controlado por Custo Médio Ponderado:
    - compra: soma o custo da compra ao custo total em carteira e recalcula o preço médio;
    - venda: baixa a quantidade vendida pelo preço médio vigente do ticker.

    NOVO: O resultado de swing trade é separado em dois acumuladores:
    - result_swing_on_pn   : lucro/prejuízo de ON e PN (elegíveis à isenção de 20k)
    - result_swing_outros  : lucro/prejuízo de BDR, UNITS, ETF (sem isenção, sempre B15)

    Também acumula separadamente:
    - vendas_on_pn  : total de vendas de ON+PN no mês (para regra dos 20k)
    - vendas_outros : total de vendas dos demais tipos swing (para somar em G20)

    A decisão B15 vs J20 para ON+PN é feita em update_month_sheet(),
    que recebe vendas_on_pn e compara com LIMITE_ISENCAO_ON_PN. O G20 recebe
    vendas_on_pn + vendas_outros, mas esse total não altera a regra dos 20k.
    """
    ops_sorted = sorted(ops, key=lambda o: (o.ref_date, o.note_file))

    # Garante uma única posição consolidada por ticker, mesmo que a origem tenha lotes.
    for ticker, lots in list(positions.items()):
        total_qty = sum(q for q, _ in lots)
        if total_qty == 0:
            positions[ticker] = []
            continue
        total_cost = sum(q * c for q, c in lots)
        avg_cost = total_cost / total_qty
        positions[ticker] = [(total_qty, avg_cost)]

    # Resultados swing separados por elegibilidade à isenção
    result_swing_on_pn  = Decimal(0)  # ON e PN
    result_swing_outros = Decimal(0)  # BDR, UNITS, ETF e tipos desconhecidos

    result_day = Decimal(0)
    result_fii = Decimal(0)

    irrf_swing = Decimal(0)
    irrf_day   = Decimal(0)
    irrf_fii   = Decimal(0)

    # Acumuladores de volume de vendas para cálculo de isenção e G20
    vendas_on_pn  = Decimal(0)  # vendas de ON+PN; usado para a regra dos 20k
    vendas_outros = Decimal(0)  # vendas de BDR/UNITS/ETF e outros swing sem isenção

    day_groups: Dict[Tuple[date, str, str], List[Operation]] = {}

    for op in ops_sorted:
        if op.category not in ("day", "swing", "fii"):
            raise RuntimeError(
                f"Operação com categoria inválida '{op.category}' na nota '{op.note_file}'."
            )

        if op.category == "day":
            key = (op.ref_date, op.ticker.strip().upper(), op.note_file)
            day_groups.setdefault(key, []).append(op)
            continue

        ticker = normalize_b3_ticker(op.ticker) or ""

        if op.category == "fii":
            irrf_fii += (op.irrf or Decimal(0))
        elif op.category == "swing":
            irrf_swing += (op.irrf or Decimal(0))

        if op.transaction_type == "buy":
            purchase_cost = (op.unit_price * op.amount) + op.allocated_fee
            current_qty, current_avg = (
                positions.get(ticker, [(Decimal(0), Decimal(0))])[0]
                if positions.get(ticker)
                else (Decimal(0), Decimal(0))
            )
            new_qty = current_qty + op.amount

            if new_qty > 0:
                if current_qty > 0:
                    current_cost = current_qty * current_avg
                    new_avg = (current_cost + purchase_cost) / new_qty
                else:
                    # Se havia posição negativa, o custo médio passa a refletir
                    # apenas a sobra positiva coberta por esta compra.
                    new_avg = purchase_cost / op.amount
                positions[ticker] = [(new_qty, new_avg)]
            elif new_qty < 0:
                positions[ticker] = [(new_qty, Decimal("0"))]
            else:
                positions[ticker] = []

        elif op.transaction_type == "sell":
            sale_value  = op.unit_price * op.amount
            current_qty, current_avg = (
                positions.get(ticker, [(Decimal(0), Decimal(0))])[0]
                if positions.get(ticker)
                else (Decimal(0), Decimal(0))
            )

            qty_available = current_qty if current_qty > 0 else Decimal(0)
            qty_costed = min(op.amount, qty_available)
            cost_total = qty_costed * current_avg
            new_qty = current_qty - op.amount

            if new_qty > 0:
                positions[ticker] = [(new_qty, current_avg)]
            elif new_qty < 0:
                positions[ticker] = [(new_qty, Decimal("0"))]
            else:
                positions[ticker] = []

            profit = sale_value - cost_total - op.allocated_fee

            if op.category == "fii":
                result_fii += profit

            elif op.category == "swing":
                # Separa o resultado pelo tipo do ativo para a regra de isenção
                tipo = op.asset_type or ""

                if tipo in ASSET_TYPES_COM_ISENCAO:
                    # ON ou PN: acumula no resultado elegível à isenção
                    result_swing_on_pn += profit
                    vendas_on_pn       += sale_value  # volume de venda para cálculo de 20k
                else:
                    # BDR, UNITS, ETF, ou tipo desconhecido: sempre tributável
                    result_swing_outros += profit
                    vendas_outros       += sale_value

    # Processa Day Trade
    for _, ops_day in day_groups.items():
        buy_total = sell_total = buy_fee = sell_fee = irrf_day_group = Decimal(0)

        for op in ops_day:
            irrf_day_group += (op.irrf or Decimal(0))
            if op.transaction_type == "buy":
                buy_total += op.total_value
                buy_fee   += op.allocated_fee
            elif op.transaction_type == "sell":
                sell_total += op.total_value
                sell_fee   += op.allocated_fee

        profit = (sell_total - sell_fee) - (buy_total + buy_fee)
        result_day += profit

        if irrf_day_group > 0:
            irrf_day += irrf_day_group
        elif profit > 0:
            irrf_day += profit * Decimal("0.01")

    irrf_day += irrf_day_from_notes

    has_negative_positions = any(
        sum(q for q, _ in lots) < 0
        for lots in positions.values()
    )

    return (
        result_swing_on_pn, result_swing_outros,
        result_day, result_fii,
        irrf_swing, irrf_day, irrf_fii,
        vendas_on_pn, vendas_outros,
        positions, has_negative_positions,
    )


###############################################################################
# Utilitários de formatação e validação
###############################################################################

def fmt(value: Decimal) -> str:
    """Formata Decimal com duas casas, usando vírgula como separador decimal."""
    return f"{value:.2f}".replace(".", ",")


def is_file_locked_for_write(path: str) -> bool:
    """Retorna True quando o arquivo não pode ser aberto para escrita."""
    try:
        with open(path, "r+b"):
            return False
    except PermissionError:
        return True
    except OSError as exc:
        raise RuntimeError(
            f"Não foi possível acessar o arquivo Excel selecionado: {exc}"
        ) from exc


def validar_cpf(cpf: str) -> bool:
    """Valida CPF (dígitos verificadores)."""
    cpf = re.sub(r"[^\d]", "", cpf)
    if len(cpf) != 11 or cpf == cpf[0] * 11:
        return False
    soma1 = sum(int(cpf[i]) * (10 - i) for i in range(9))
    dig1  = (soma1 * 10 % 11) % 10
    if dig1 != int(cpf[9]):
        return False
    soma2 = sum(int(cpf[i]) * (11 - i) for i in range(10))
    dig2  = (soma2 * 10 % 11) % 10
    return dig2 == int(cpf[10])


def formatar_cpf(cpf: str) -> str:
    """Formata CPF no padrão XXX.XXX.XXX-XX."""
    cpf = re.sub(r"[^\d]", "", cpf)
    return f"{cpf[:3]}.{cpf[3:6]}.{cpf[6:9]}-{cpf[9:]}"


###############################################################################
# Exportação de Debug para Excel (SOMENTE sob demanda do usuário)
###############################################################################

def export_debug_excel(operations: List[Operation], output_path: str) -> None:
    """Gera um Excel com todas as operações lidas para conferência.

    Gerado APENAS quando o usuário clicar em 'Exportar Debug'.
    Inclui a coluna 'Tipo do Ativo' para facilitar a verificação da detecção automática.
    """
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Memória de Cálculo de Operações"

    headers = [
        "#", "Data", "Ticker", "Nome", "Tipo Ativo",
        "Operação", "Quantidade", "Preço Unitário", "Valor Total",
        "Taxa Rateada", "IRRF", "Categoria", "Arquivo PDF", "Parser Utilizado",
    ]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font  = header_font
        cell.fill  = header_fill
        cell.alignment = Alignment(horizontal="center")

    for i, op in enumerate(operations, start=1):
        ws.append([
            i,
            op.ref_date.strftime("%d/%m/%Y"),
            op.ticker or "",
            op.name,
            op.asset_type or "?",           
            op.transaction_type,
            float(op.amount),
            float(op.unit_price),
            float(op.total_value),
            float(op.allocated_fee),
            float(op.irrf) if op.irrf else 0.0,
            op.category or "",
            op.note_file,
            op.parser_used,
        ])

    for col in ws.columns:
        max_len = max((len(str(cell.value)) for cell in col if cell.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    wb.save(output_path)


###############################################################################
# Funções de planilha Excel (resultado da apuração)
###############################################################################

def resolve_month_sheet_name(workbook, expected_month: int, expected_year: int) -> str:
    """Encontra a aba mensal aceitando MM.AA, MM.AAAA, MM/AA ou MM/AAAA."""
    matches: List[str] = []
    pattern = re.compile(r"^\s*(\d{1,2})\s*[./]\s*(\d{2}|\d{4})\s*$")

    for sheet_name in workbook.sheetnames:
        match = pattern.fullmatch(str(sheet_name))
        if not match:
            continue
        month     = int(match.group(1))
        year_text = match.group(2)
        if month != expected_month:
            continue
        if len(year_text) == 2:
            if int(year_text) == expected_year % 100:
                matches.append(sheet_name)
        elif int(year_text) == expected_year:
            matches.append(sheet_name)

    if not matches:
        raise RuntimeError(
            f"Nenhuma aba do mês {expected_month:02d}/{expected_year} foi encontrada na planilha de apuração. "
            "Use um destes padrões: MM/AA ou MM/AAAA."
        )
    if len(matches) > 1:
        raise RuntimeError(
            f"Foram encontradas várias abas para {expected_month:02d}/{expected_year}: "
            f"{', '.join(matches)}. Deixe apenas uma aba mensal correspondente."
        )
    return matches[0]


def update_month_sheet(
    workbook,
    sheet_name: str,
    name: str,
    cpf: str,
    result_swing_on_pn: Decimal,    # resultado de ON+PN (pode ir para J20 ou B15)
    result_swing_outros: Decimal,   # resultado de BDR/UNITS/ETF (sempre B15)
    result_day: Decimal,
    result_fii: Decimal,
    irrf_swing: Decimal,
    irrf_day: Decimal,
    irrf_fii: Decimal,
    vendas_on_pn: Decimal,          # volume de vendas de ON+PN no mês; define a isenção de 20k
    vendas_outros: Decimal,         # volume de vendas dos outros tipos swing; soma em G20
):
    """Atualiza a aba do mês com os resultados da apuração.

    LÓGICA DE PREENCHIMENTO PARA SWING TRADE:
    ┌─────────────────────────────────────────────────────────────────┐
    │ Tipo    │ Condição           │ Onde grava o resultado           │
    ├─────────┼────────────────────┼──────────────────────────────────┤
    │ ON + PN │ Vendas < 20k       │ J20 (ganho isento)               │
    │ ON + PN │ Vendas ≥ 20k       │ B15 (tributável)                 │
    │ BDR     │ sempre             │ B15 (tributável)                 │
    │ UNITS   │ sempre             │ B15 (tributável)                 │
    │ ETF     │ sempre             │ B15 (tributável)                 │
    └─────────┴────────────────────┴──────────────────────────────────┘

    G20 recebe o total de alienações do mês para swing:
    ON + PN + BDR + UNITS + ETF + demais tipos sem isenção.

    A regra de isenção de 20k, porém, compara apenas vendas_on_pn.
    """
    sheet = workbook[sheet_name]

    sheet["A6"] = "PERÍODO DE APURAÇÃO: " + sheet_name.strip()
    sheet["A7"] = "INVESTIDOR: " + name
    sheet["A8"] = "CPF: " + formatar_cpf(cpf)

    # ── Day Trade ──────────────────────────────────────────────────────────────
    sheet["C15"] = float(result_day)
    sheet["C36"] = float(irrf_day)

    # ── FII ───────────────────────────────────────────────────────────────────
    sheet["B49"] = float(result_fii)
    sheet["B56"] = float(irrf_fii)

    # ── IRRF Swing ────────────────────────────────────────────────────────────
    sheet["C39"] = float(irrf_swing)

    # ── Swing Trade — ON e PN (lógica de isenção 20k) ─────────────────────────
    # G20 mostra o total vendido no mês em swing, incluindo ativos sem isenção.
    # A decisão de isenção continua usando somente vendas_on_pn.
    sheet["G20"] = float(vendas_on_pn + vendas_outros)

    if vendas_on_pn < LIMITE_ISENCAO_ON_PN:
        # Abaixo de 20k de vendas: lucro de ON+PN vai para J20 (ganho isento)
        # B15 recebe APENAS o resultado dos outros tipos (BDR, UNITS, ETF)
        sheet["J20"] = float(result_swing_on_pn)
        sheet["B15"] = float(result_swing_outros)
    else:
        # Acima de 20k de vendas: lucro de ON+PN também vai para B15 (tributável)
        # J20 fica zerado (não há ganho isento)
        sheet["J20"] = 0.0
        sheet["B15"] = float(result_swing_on_pn + result_swing_outros)


def _to_decimal(value) -> Optional[Decimal]:
    """Converte valor lido do Excel em Decimal."""
    if value is None or value == "":
        return None
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))
    if isinstance(value, str):
        txt = value.strip()
        if not txt:
            return None
        if "," in txt:
            txt = txt.replace(".", "").replace(",", ".")
        try:
            return Decimal(txt)
        except Exception:
            return None
    return None


def read_initial_portfolio_from_year_sheet(
    workbook,
    year_sheet_name: str,
) -> Dict[str, List[Tuple[Decimal, Decimal]]]:
    """Lê as carteiras do mês anterior na aba anual.

    Cada carteira deve ter o título exato 'ATIVOS EM CARTEIRA' com colunas:
        PAPEL | QUANTIDADE | PREÇO MÉDIO | CUSTO TOTAL
    """
    if year_sheet_name not in workbook.sheetnames:
        raise RuntimeError(
            f"A aba anual '{year_sheet_name}' não foi encontrada na planilha."
        )

    sheet = workbook[year_sheet_name]
    positions: Dict[str, List[Tuple[Decimal, Decimal]]] = {}
    found_valid_table = False

    for title_row in range(1, sheet.max_row + 1):
        for title_col in range(1, sheet.max_column + 1):
            title = sheet.cell(row=title_row, column=title_col).value
            if not isinstance(title, str) or title.strip().lower() != "ativos em carteira":
                continue

            columns_row = title_row + 1
            papel = sheet.cell(row=columns_row, column=title_col).value
            quantidade = sheet.cell(row=columns_row, column=title_col + 1).value
            preco_medio = sheet.cell(row=columns_row, column=title_col + 2).value

            if (
                not isinstance(papel, str)
                or papel.strip().lower() != "papel"
                or not isinstance(quantidade, str)
                or quantidade.strip().lower() != "quantidade"
                or not isinstance(preco_medio, str)
                or preco_medio.strip().lower() not in ("preço médio", "preco medio")
            ):
                continue

            found_valid_table = True
            row = columns_row + 1

            while row <= sheet.max_row:
                ticker = sheet.cell(row=row, column=title_col).value
                if ticker is None or (isinstance(ticker, str) and not ticker.strip()):
                    break
                if isinstance(ticker, str) and "valor total de patrimônio acumulado" in ticker.lower():
                    break

                ticker_str = normalize_b3_ticker(str(ticker)) or ""
                qty = _to_decimal(sheet.cell(row=row, column=title_col + 1).value)
                avg_price = _to_decimal(sheet.cell(row=row, column=title_col + 2).value)

                if ticker_str and qty is not None and avg_price is not None and qty != 0:
                    positions.setdefault(ticker_str, []).append((qty, avg_price))

                row += 1

    if not found_valid_table:
        raise RuntimeError(
            f"Tabela válida 'ATIVOS EM CARTEIRA' não encontrada na aba '{year_sheet_name}'."
        )

    return positions


def write_portfolio_sheet_for_month(
    workbook,
    month_sheet_name: str,
    positions: Dict[str, List[Tuple[Decimal, Decimal]]],
):
    """Cria/atualiza a aba 'CARTEIRA_MM.AA' com a carteira final do mês."""
    carteira_sheet_name = f"CARTEIRA_{month_sheet_name}"
    if carteira_sheet_name in workbook.sheetnames:
        del workbook[carteira_sheet_name]

    sheet          = workbook.create_sheet(carteira_sheet_name)
    sheet["A1"]    = "PAPEL"
    sheet["B1"]    = "QUANTIDADE"
    sheet["C1"]    = "PREÇO MÉDIO"
    sheet["D1"]    = "CUSTO TOTAL"

    row = 2
    for ticker, lots in positions.items():
        total_qty = sum(q for q, _ in lots)
        if total_qty == 0:
            continue
        total_cost = sum(q * c for q, c in lots)
        avg_price  = total_cost / total_qty

        sheet.cell(row=row, column=1).value = ticker
        sheet.cell(row=row, column=2).value = float(total_qty)
        sheet.cell(row=row, column=3).value = float(avg_price)
        sheet.cell(row=row, column=4).value = float(total_cost)
        row += 1


###############################################################################
# Interface Gráfica (Tkinter)
###############################################################################

class ApuracaoB3App:
    """Classe principal da aplicação gráfica Tkinter."""

    def __init__(self, master: tk.Tk, base_dir: str):
        self.master   = master
        self.base_dir = base_dir

        self.master.title("Calculadora B3 - Notas de Corretagem                                             v 1.1")
        W, H = 550, 390
        self.master.geometry(f"{W}x{H}")
        self.master.resizable(False, False)
        self.master.iconbitmap(default=os.path.join(self.base_dir, "icon.ico"))

        self.canvas = tk.Canvas(self.master, width=W, height=H, highlightthickness=0, bd=0)
        self.canvas.pack(fill="both", expand=True)
        self._set_background_image("background.png")

        self.investor_name        = tk.StringVar()
        self.cpf                  = tk.StringVar()
        self.selected_month_folder = tk.StringVar()
        self.selected_excel_path  = tk.StringVar()

        # Mapa de senhas por arquivo: {filename: senha}
        self.password_map: Dict[str, Optional[str]] = {}

        # Mapa de tickers persistente: {nome_do_ativo: ticker}
        self.ticker_map_path = get_data_path("ticker_map.json")
        self.ticker_map: Dict[str, str] = load_json_dict(self.ticker_map_path)

        # Mapa de tipos de ativo persistente: {ticker: tipo} (ON, PN, BDR, etc.)
        # Evita perguntar o mesmo ativo duas vezes em apurações futuras.
        self.asset_type_map_path = get_data_path("asset_type_map.json")
        self.asset_type_map: Dict[str, str] = load_json_dict(self.asset_type_map_path)

        # Últimas operações processadas (para exportar debug)
        self._last_operations: Optional[List[Operation]] = None

        self._build_widgets()

    def _set_background_image(self, image_filename: str):
        """Define a imagem de fundo da janela."""
        image_path = resource_path(image_filename)
        if not os.path.isfile(image_path):
            return
        self.master.update_idletasks()
        w = self.master.winfo_width()
        h = self.master.winfo_height()
        img = Image.open(image_path).resize((w, h), Image.LANCZOS)
        self.bg_image = ImageTk.PhotoImage(img)
        self.canvas.create_image(0, 0, image=self.bg_image, anchor="nw")

    def _build_widgets(self):
        """Constrói todos os widgets da interface gráfica."""
        x_label = 20
        x_entry = 200
        y       = 55
        gap     = 30

        tk.Label(self.master, text="Nome do investidor:", width=23, height=1, font=("Segoe UI", 9, "bold")).place(x=x_label, y=y)
        tk.Entry(self.master, textvariable=self.investor_name, width=54).place(x=x_entry, y=y)

        y += gap
        tk.Label(self.master, text="CPF do investidor:", width=23, height=1, font=("Segoe UI", 9, "bold")).place(x=x_label, y=y)
        tk.Entry(self.master, textvariable=self.cpf, width=54).place(x=x_entry, y=y)

        y += gap + 6
        tk.Button(self.master, text="Selecione a pasta do mês", command=self._select_folder, width=23, height=1, font=("Segoe UI", 9, "bold")).place(x=x_label, y=y)
        tk.Entry(self.master, textvariable=self.selected_month_folder, width=54).place(x=x_entry, y=y)

        y += gap + 6
        tk.Button(self.master, text="Selecione o Excel da apuração", command=self._select_excel, width=23, height=1, font=("Segoe UI", 9, "bold")).place(x=x_label, y=y)
        tk.Entry(self.master, textvariable=self.selected_excel_path, width=54).place(x=x_entry, y=y)

        y += gap + 20
        tk.Button(self.master, text="Calcular Apuração", command=self._run_apuracao, width=71, font=("Segoe UI", 9, "bold")).place(x=x_label, y=y)

        # Botão de debug no canto inferior direito; desabilitado até haver apuração
        self.btn_debug = tk.Button(self.master, text="Exportar Memória de Cálculo", command=self._export_debug, width=21, font=("Segoe UI", 8, "bold"), state="disabled",
        )
        self.btn_debug.place(relx=1.0, rely=1.0, anchor="se", x=-15, y=-10)

        y += gap + 14
        self.status_label = tk.Label(
            self.master,
            text="Preencha todos os campos antes de calcular.",
            font=("Segoe UI", 9, "bold"),
        )
        self.status_label.place(x=x_label, y=y)

    # ── Seleção de arquivos/pastas ─────────────────────────────────────────────

    def _select_folder(self):
        folder = filedialog.askdirectory(title="Selecione a pasta do mês da apuração")
        if folder:
            self.selected_month_folder.set(folder)

    def _select_excel(self):
        file_path = filedialog.askopenfilename(
            title="Selecione a planilha Excel da apuração",
            filetypes=[("Arquivos Excel", "*.xlsx *.xlsm"), ("Todos os arquivos", "*.*")],
        )
        if file_path:
            self.selected_excel_path.set(file_path)

    # ── Janelas modais ─────────────────────────────────────────────────────────

    def _pedir_senha_pdf(self, filename: str) -> Optional[str]:
        """Modal para solicitar a senha de um PDF específico."""
        win = tk.Toplevel(self.master)
        win.title("Senha do PDF")
        win.grab_set()

        tk.Label(
            win,
            text=f"O arquivo abaixo está protegido por senha:\n\n{filename}\n\nDigite a senha:",
            justify="center",
        ).pack(padx=10, pady=8)

        senha_var = tk.StringVar()
        tk.Entry(win, textvariable=senha_var, show="*", width=50).pack(padx=10, pady=5)

        result = {"senha": None, "cancelou": False}

        def confirmar():
            result["senha"] = senha_var.get().strip() or None
            win.destroy()

        def cancelar():
            result["cancelou"] = True
            win.destroy()

        btn_frame = tk.Frame(win)
        btn_frame.pack(pady=10)
        tk.Button(btn_frame, text="OK",      command=confirmar, width=12).pack(side="left",  padx=5)
        tk.Button(btn_frame, text="Cancelar", command=cancelar,  width=12).pack(side="right", padx=5)

        self.master.wait_window(win)
        return None if result["cancelou"] else result["senha"]

    def _pedir_quantidade_trade(
        self,
        note_file: str,
        ref_date: date,
        ticker: str,
        tx_type: str,
        unit_price: Decimal,
    ) -> Optional[Decimal]:
        """Modal para solicitar quantidade de um trade quando não identificada automaticamente."""
        win = tk.Toplevel(self.master)
        win.title("Quantidade necessária")
        win.grab_set()

        msg = (
            "Não foi possível obter a quantidade desta operação automaticamente.\n\n"
            f"Arquivo: {note_file}\n"
            f"Data: {ref_date:%d/%m/%Y}\n"
            f"Ticker: {ticker}\n"
            f"Tipo: {tx_type}\n"
            f"Preço: {unit_price:.2f}\n\n"
            "Informe a quantidade (apenas o número inteiro):"
        )
        tk.Label(win, text=msg, justify="left", anchor="w").pack(padx=10, pady=10)

        qty_var = tk.StringVar()
        entry = tk.Entry(win, textvariable=qty_var, width=30)
        entry.pack(padx=10, pady=5)
        entry.focus_set()

        result = {"qty": None}

        def confirmar():
            txt = (qty_var.get() or "").strip().replace(".", "")
            if not txt or not txt.isdigit():
                messagebox.showwarning("Valor inválido", "Digite apenas números.")
                return
            q = int(txt)
            if q <= 0:
                messagebox.showwarning("Valor inválido", "A quantidade deve ser maior que zero.")
                return
            result["qty"] = Decimal(q)
            win.destroy()

        def cancelar():
            win.destroy()

        btn = tk.Frame(win)
        btn.pack(padx=10, pady=10, fill="x")
        tk.Button(btn, text="OK",       command=confirmar).pack(side="right")
        tk.Button(btn, text="Cancelar", command=cancelar).pack(side="left")
        self.master.wait_window(win)
        return result["qty"]

    def _pedir_ticker(self, note_file: str, ref_date: date, name: str) -> Optional[str]:
        """Modal para solicitar o ticker de um ativo não identificado automaticamente."""
        win = tk.Toplevel(self.master)
        win.title("Ticker necessário")
        win.grab_set()

        tk.Label(
            win,
            text=(
                "Não foi possível identificar o TICKER automaticamente.\n\n"
                f"Arquivo: {note_file}\n"
                f"Data: {ref_date:%d/%m/%Y}"
            ),
            justify="left",
            anchor="w",
        ).pack(padx=10, pady=(10, 4), anchor="w")

        tk.Label(win, text="Ativo (nome na nota):", anchor="w").pack(
            padx=10, pady=(4, 0), anchor="w"
        )
        name_var = tk.StringVar(value=name)
        name_entry = tk.Entry(win, textvariable=name_var, width=50, state="readonly")
        name_entry.pack(padx=10, pady=(0, 10), fill="x")
        name_entry.bind("<Control-a>", lambda event: (name_entry.select_range(0, "end"), "break"))

        tk.Label(
            win,
            text="Digite o ticker (ex.: PETR4, VALE3, HGLG11):",
            anchor="w",
        ).pack(padx=10, pady=(0, 5), anchor="w")

        ticker_var = tk.StringVar()
        entry = tk.Entry(win, textvariable=ticker_var, width=30)
        entry.pack(padx=10, pady=5)
        entry.focus_set()

        result = {"ticker": None}

        def confirmar():
            t = (ticker_var.get() or "").strip().upper()
            if not t:
                messagebox.showwarning("Campo obrigatório", "O ticker não pode ficar vazio.")
                return
            if not re.fullmatch(r"[A-Z0-9]{4,10}", t):
                messagebox.showwarning("Ticker inválido", "Use apenas letras e números (ex.: PETR4, HGLG11).")
                return
            result["ticker"] = t
            win.destroy()

        def cancelar():
            win.destroy()

        btn = tk.Frame(win)
        btn.pack(padx=10, pady=10, fill="x")
        tk.Button(btn, text="OK",       command=confirmar).pack(side="right")
        tk.Button(btn, text="Cancelar", command=cancelar).pack(side="left")
        self.master.wait_window(win)
        return result["ticker"]

    def _pedir_tipo_ativo(
        self,
        ticker: str,
        name: str,
        note_file: str,
        ref_date: date,
    ) -> Optional[str]:
        """Modal para solicitar o tipo do ativo quando a detecção automática falha.

        Exibe botões para cada tipo possível (ON, PN, FII, FIAGRO, BDR, UNITS, ETF).
        O tipo escolhido é salvo no asset_type_map (JSON) para não perguntar novamente.

        Retorna o identificador canônico do tipo (ex.: "ON"), ou None se o usuário cancelar.
        """
        win = tk.Toplevel(self.master)
        win.title("Tipo do Ativo")
        win.grab_set()
        win.resizable(False, False)

        msg = (
            "Não foi possível identificar automaticamente o tipo do ativo abaixo.\n\n"
            f"Ticker : {ticker}\n"
            f"Nome   : {name}\n"
            f"Arquivo: {note_file}\n"
            f"Data   : {ref_date:%d/%m/%Y}\n\n"
            "Selecione o tipo correto:"
        )
        tk.Label(win, text=msg, justify="left", anchor="w").pack(padx=15, pady=10)

        # Descrições amigáveis para cada tipo
        tipos = [
            (ASSET_TYPE_ON,     "ON  — Ordinária  (tem isenção 20k)"),
            (ASSET_TYPE_PN,     "PN  — Preferencial  (tem isenção 20k)"),
            (ASSET_TYPE_FII,    "FII — Fundo Imobiliário  (sem isenção)"),
            (ASSET_TYPE_FIAGRO, "FIAGRO — Fundo Agroindustrial  (sem isenção)"),
            (ASSET_TYPE_BDR,    "BDR — Recibo negociável  (sem isenção)"),
            (ASSET_TYPE_UNITS,  "UNITS — Units / UNT  (sem isenção)"),
            (ASSET_TYPE_ETF,    "ETF — Fundo de índice  (sem isenção)"),
        ]

        result = {"tipo": None}

        btn_frame = tk.Frame(win)
        btn_frame.pack(padx=15, pady=5, fill="x")

        for tipo_id, descricao in tipos:
            def on_click(t=tipo_id):
                result["tipo"] = t
                win.destroy()
            tk.Button(
                btn_frame,
                text=descricao,
                command=on_click,
                width=48,
                anchor="w",
                font=("Segoe UI", 9),
            ).pack(pady=2)

        def cancelar():
            win.destroy()

        tk.Button(win, text="Cancelar apuração", command=cancelar,
                  font=("Segoe UI", 9), fg="red").pack(pady=8)

        self.master.wait_window(win)
        return result["tipo"]

    def _revisao_manual_nota(
        self,
        filename: str,
        raw_text: str,
        errors: List[str],
    ) -> Optional[List[dict]]:
        """Tela de revisão manual quando todas as camadas automáticas falham.

        Exibe uma tabela visual com uma coluna por campo e uma linha por operação.
        O usuário preenche as células diretamente, sem precisar decorar um formato
        de texto separado por ponto-e-vírgula.

        Colunas da tabela:
            Data (DD/MM/AAAA) | C/V | Ticker | Nome | Quantidade | Preço | Total (opcional)

        Botões:
            [+ Linha]          → adiciona uma linha em branco ao final da tabela
            [✕] por linha      → remove aquela linha específica
            [Confirmar]        → valida e retorna as linhas preenchidas
            [Cancelar nota]    → cancela o processamento desta nota
        """
        win = tk.Toplevel(self.master)
        win.title("Revisão Manual da Nota")
        win.grab_set()
        win.geometry("980x700")
        win.resizable(True, True)

        result = {"rows": None}

        # ── Cabeçalho informativo ──────────────────────────────────────────────
        tk.Label(
            win,
            text=(
                f"Nenhum parser conseguiu ler esta nota automaticamente.\n"
                f"Arquivo: {filename}\n\n"
                "Preencha as operações na tabela abaixo e informe as taxas totais da nota."
            ),
            justify="left",
            anchor="w",
            font=("Segoe UI", 9, "bold"),
        ).pack(fill="x", padx=10, pady=(10, 4))

        # ── Erros das camadas automáticas ─────────────────────────────────────
        error_text = "\n".join(f"• {e}" for e in errors[-6:])
        tk.Label(win, text="Erros das camadas automáticas:", anchor="w",
                 font=("Segoe UI", 8)).pack(fill="x", padx=10)
        err_box = tk.Text(win, height=4, wrap="word", font=("Segoe UI", 8),
                          bg="#fff8f0", relief="flat", bd=1)
        err_box.pack(fill="x", padx=10, pady=(0, 6))
        err_box.insert("1.0", error_text or "(sem mensagens de erro)")
        err_box.config(state="disabled")

        # ── Taxas totais da nota ───────────────────────────────────────────────
        taxas_frame = tk.LabelFrame(win, text="Taxas totais da nota  (deixe 0,00 se não houver)")
        taxas_frame.pack(fill="x", padx=10, pady=(0, 8))

        taxa_liq_var = tk.StringVar(value="0,00")
        emol_var     = tk.StringVar(value="0,00")
        reg_var      = tk.StringVar(value="0,00")
        transf_var   = tk.StringVar(value="0,00")
        ana_var      = tk.StringVar(value="0,00")
        outras_var   = tk.StringVar(value="0,00")
        irrf_var     = tk.StringVar(value="0,00")

        taxa_campos = [
            ("Taxa liquidação (R$):", taxa_liq_var),
            ("Taxa registro (R$):",   reg_var),
            ("Taxa A.N.A. (R$):",     ana_var),
            ("Emolumentos (R$):",     emol_var),
            ("Taxa transferência (R$):", transf_var),
            ("Outros (R$):",          outras_var),
            ("IRRF (R$):",            irrf_var),
        ]
        for col_idx, (lbl, var) in enumerate(taxa_campos):
            row_idx = col_idx // 4
            pair_col = (col_idx % 4) * 2
            tk.Label(taxas_frame, text=lbl, font=("Segoe UI", 8)).grid(
                row=row_idx, column=pair_col, sticky="e", padx=(8, 2), pady=4)
            tk.Entry(taxas_frame, textvariable=var, width=10,
                     font=("Segoe UI", 8)).grid(
                row=row_idx, column=pair_col + 1, sticky="w", padx=(0, 8), pady=4)

        # ── Tabela de operações ────────────────────────────────────────────────
        # Estrutura: frame com scroll vertical.
        # Cada linha é uma lista de tk.StringVar + tk.Entry + botão [✕].

        table_outer = tk.LabelFrame(win, text="Operações  (uma linha por operação)")
        table_outer.pack(fill="both", expand=True, padx=10, pady=(0, 6))

        # Canvas com scrollbar para suportar muitas linhas
        canvas = tk.Canvas(table_outer, borderwidth=0, highlightthickness=0)
        vsb    = tk.Scrollbar(table_outer, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        # Frame interno onde as linhas são colocadas
        inner = tk.Frame(canvas)
        canvas_window = canvas.create_window((0, 0), window=inner, anchor="nw")

        def _on_inner_configure(event):
            canvas.configure(scrollregion=canvas.bbox("all"))
            # Faz o inner ter sempre a largura do canvas
            canvas.itemconfig(canvas_window, width=canvas.winfo_width())

        def _on_canvas_configure(event):
            canvas.itemconfig(canvas_window, width=event.width)

        inner.bind("<Configure>", _on_inner_configure)
        canvas.bind("<Configure>", _on_canvas_configure)

        # Scroll com mouse wheel
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)

        # Definição das colunas: (chave_interna, texto_cabeçalho, largura_entry)
        COLUNAS = [
            ("data",       "Data\n(DD/MM/AAAA)", 12),
            ("cv",         "C/V\n(C=compra\nV=venda)", 5),
            ("ticker",     "Ticker\n(ex: PETR4)", 9),
            ("nome",       "Nome do ativo\n(ex: PETROBRAS ON)", 22),
            ("quantidade", "Quantidade\n(ex: 100)", 10),
            ("preco",      "Preço unit.\n(ex: 28,50)", 10),
            ("total",      "Total R$\n(opcional)", 10),
        ]

        # Linha de cabeçalho da tabela
        for col_idx, (_, header_text, _) in enumerate(COLUNAS):
            tk.Label(
                inner,
                text=header_text,
                font=("Segoe UI", 8, "bold"),
                bg="#1f4e79",
                fg="white",
                relief="flat",
                anchor="center",
                justify="center",
                padx=4, pady=4,
            ).grid(row=0, column=col_idx, sticky="nsew", padx=1, pady=(2, 1))

        # Cabeçalho da coluna do botão remover
        tk.Label(
            inner, text=" ", bg="#1f4e79", fg="white", width=3,
        ).grid(row=0, column=len(COLUNAS), padx=1, pady=(2, 1))

        # Configura pesos das colunas para que a coluna "nome" expanda
        for col_idx in range(len(COLUNAS)):
            inner.columnconfigure(col_idx, weight=1 if col_idx == 3 else 0)

        # Lista de linhas: cada elemento é um dict {chave: StringVar}
        linhas: List[Dict[str, tk.StringVar]] = []

        def _adicionar_linha(dados: Optional[Dict[str, str]] = None) -> None:
            """Adiciona uma nova linha de entrada na tabela.

            Se 'dados' for fornecido, preenche as células com os valores.
            """
            row_idx = len(linhas) + 1   # +1 porque a linha 0 é o cabeçalho
            vars_linha: Dict[str, tk.StringVar] = {}

            # Cor de fundo alternada para facilitar leitura visual
            bg_color = "#f5f9ff" if row_idx % 2 == 0 else "#ffffff"

            for col_idx, (key, _, width) in enumerate(COLUNAS):
                var = tk.StringVar(value=(dados or {}).get(key, ""))
                vars_linha[key] = var

                entry = tk.Entry(
                    inner,
                    textvariable=var,
                    width=width,
                    font=("Segoe UI", 9),
                    bg=bg_color,
                    relief="solid",
                    bd=1,
                )
                entry.grid(row=row_idx, column=col_idx, sticky="ew", padx=1, pady=1)

                # Tab para avançar naturalmente entre campos
                entry.bind("<Tab>", lambda e, w=entry: w.tk_focusNext().focus())

            # Botão para remover esta linha
            # Usa uma closure para capturar o índice correto
            def _remover(idx=len(linhas)):
                _remover_linha(idx)

            btn_rem = tk.Button(
                inner,
                text="✕",
                font=("Segoe UI", 8),
                fg="#c00000",
                relief="flat",
                padx=2,
                command=_remover,
            )
            btn_rem.grid(row=row_idx, column=len(COLUNAS), padx=2, pady=1)

            # Salva o botão junto com as vars para poder reposicioná-lo depois
            vars_linha["__btn_rem__"] = btn_rem
            vars_linha["__row_widgets__"] = []   # widgets Entry desta linha

            # Coleta os Entry widgets para poder removê-los depois
            for col_idx in range(len(COLUNAS)):
                widget = inner.grid_slaves(row=row_idx, column=col_idx)
                if widget:
                    vars_linha["__row_widgets__"].append(widget[0])

            linhas.append(vars_linha)

            # Atualiza scroll
            inner.update_idletasks()
            canvas.configure(scrollregion=canvas.bbox("all"))
            canvas.yview_moveto(1.0)   # rola para o final ao adicionar nova linha

        def _remover_linha(idx: int) -> None:
            """Remove a linha de índice 'idx' da tabela e reconstrói o layout."""
            if idx >= len(linhas):
                return

            # Destrói todos os widgets daquela linha
            linha = linhas[idx]
            for w in linha.get("__row_widgets__", []):
                try:
                    w.destroy()
                except Exception:
                    pass
            try:
                linha["__btn_rem__"].destroy()
            except Exception:
                pass

            linhas.pop(idx)

            # Reposiciona todas as linhas restantes para fechar o "buraco"
            for new_idx, l in enumerate(linhas):
                row_grid = new_idx + 1   # +1 por causa do cabeçalho

                # Atualiza cor de fundo alternada
                bg_color = "#f5f9ff" if row_grid % 2 == 0 else "#ffffff"

                for col_idx, (key, _, _) in enumerate(COLUNAS):
                    ws = l.get("__row_widgets__", [])
                    if col_idx < len(ws):
                        ws[col_idx].grid(row=row_grid, column=col_idx)
                        ws[col_idx].config(bg=bg_color)

                # Reposiciona e reconfigura o botão remover com o índice correto
                def _novo_remover(i=new_idx):
                    _remover_linha(i)

                l["__btn_rem__"].grid(row=row_grid, column=len(COLUNAS))
                l["__btn_rem__"].config(command=_novo_remover)

            inner.update_idletasks()
            canvas.configure(scrollregion=canvas.bbox("all"))

        # Começa com 5 linhas em branco
        for _ in range(5):
            _adicionar_linha()

        # ── Botões de ação ─────────────────────────────────────────────────────
        action_frame = tk.Frame(win)
        action_frame.pack(fill="x", padx=10, pady=6)

        tk.Button(
            action_frame,
            text="+ Adicionar linha",
            command=lambda: _adicionar_linha(),
            font=("Segoe UI", 9),
            width=18,
        ).pack(side="left", padx=5)

        # Texto extraído do PDF: fica num frame expansível no rodapé
        def _toggle_raw():
            """Mostra/esconde o texto bruto extraído do PDF."""
            if raw_frame.winfo_ismapped():
                raw_frame.pack_forget()
                btn_raw.config(text="▼ Mostrar texto do PDF")
            else:
                raw_frame.pack(fill="both", expand=False, padx=10, pady=(0, 4),
                               before=action_frame)
                btn_raw.config(text="▲ Ocultar texto do PDF")

        btn_raw = tk.Button(
            action_frame,
            text="▼ Mostrar texto do PDF",
            command=_toggle_raw,
            font=("Segoe UI", 8),
            width=22,
        )
        btn_raw.pack(side="left", padx=5)

        tk.Button(
            action_frame,
            text="Cancelar nota",
            command=lambda: (win.destroy(),),
            font=("Segoe UI", 9),
            fg="#c00000",
            width=14,
        ).pack(side="right", padx=5)

        def confirmar():
            """Coleta os valores de todas as linhas e chama _parse_manual_review_lines."""
            try:
                taxas = {
                    "settlement_fee":   _parse_decimal_br(taxa_liq_var.get()) or Decimal("0"),
                    "emoluments":       _parse_decimal_br(emol_var.get())     or Decimal("0"),
                    "registration_fee": _parse_decimal_br(reg_var.get())      or Decimal("0"),
                    "transfer_fee":     _parse_decimal_br(transf_var.get())   or Decimal("0"),
                    "ana_fee":          _parse_decimal_br(ana_var.get())      or Decimal("0"),
                    "other_fee":        _parse_decimal_br(outras_var.get())   or Decimal("0"),
                    "irrf":             _parse_decimal_br(irrf_var.get())      or Decimal("0"),
                }

                # Monta a lista de dicts a partir das StringVars de cada linha
                grid_rows = []
                for linha in linhas:
                    grid_rows.append({
                        key: linha[key].get().strip()
                        for key, _, _ in COLUNAS
                    })

                result["rows"] = _parse_manual_review_lines(grid_rows, taxas)
                win.destroy()

            except Exception as exc:
                messagebox.showerror("Erro ao confirmar operações", str(exc))

        tk.Button(
            action_frame,
            text="✔ Confirmar operações",
            command=confirmar,
            font=("Segoe UI", 9, "bold"),
            width=22,
        ).pack(side="right", padx=5)

        # Frame do texto bruto do PDF (começa oculto)
        raw_frame = tk.LabelFrame(win, text="Texto extraído do PDF (somente leitura)")
        raw_box = tk.Text(raw_frame, height=7, wrap="word", font=("Segoe UI", 8))
        raw_box.pack(fill="both", expand=True, padx=6, pady=4)
        raw_box.insert("1.0", (raw_text or "")[:12000])
        raw_box.config(state="disabled")
        # Não empacota o raw_frame ainda; _toggle_raw() faz isso sob demanda

        self.master.wait_window(win)

        # Limpa o binding do scroll de mouse ao fechar a janela
        try:
            canvas.unbind_all("<MouseWheel>")
        except Exception:
            pass

        return result["rows"]

    # ── Persistência dos mapas ─────────────────────────────────────────────────

    def _save_ticker_map(self) -> None:
        save_json_dict(self.ticker_map_path, self.ticker_map)

    def _save_asset_type_map(self) -> None:
        """Salva o mapa de tipos de ativo no disco para reutilização em sessões futuras."""
        save_json_dict(self.asset_type_map_path, self.asset_type_map)

    # ── Resolução de tipos de ativo ───────────────────────────────────────────

    def resolve_asset_types(self, operations: List[Operation]) -> None:
        """Preenche o campo asset_type de todas as operações.

        Estratégia por operação:
        1. Tenta detect_asset_type() (automático por nome + ticker).
        2. Se automático falhar, consulta o asset_type_map (JSON persistente).
        3. Se ainda não encontrar, abre o modal _pedir_tipo_ativo() para o usuário.
        4. Salva a resposta no asset_type_map para não perguntar novamente.

        Operações de day trade não precisam de tipo para o cálculo de IR
        (day trade sempre vai para C15, independente de tipo), mas preenchemos
        assim mesmo para o debug Excel ficar completo.
        """
        for op in operations:
            if op.asset_type:
                # Já preenchido (ex.: por um parser futuro mais esperto)
                continue

            ticker = normalize_b3_ticker(op.ticker) or ""

            # Tenta detecção automática
            tipo = detect_asset_type(op.name, ticker)

            if tipo is None:
                # Consulta o mapa persistente pelo ticker
                tipo = self.asset_type_map.get(ticker)

            if tipo is None:
                # Pergunta ao usuário
                tipo = self._pedir_tipo_ativo(
                    ticker=ticker,
                    name=op.name,
                    note_file=op.note_file,
                    ref_date=op.ref_date,
                )
                if tipo is None:
                    raise RuntimeError(
                        f"Apuração cancelada: tipo do ativo não informado para '{ticker}' "
                        f"({op.name}) na nota '{op.note_file}'."
                    )
                # Persiste para futuras apurações
                self.asset_type_map[ticker] = tipo
                self._save_asset_type_map()

            op.asset_type = tipo

    # ── Exportação de debug Excel ──────────────────────────────────────────────

    def _export_debug(self):
        """Exporta as operações da última apuração para um Excel de debug."""
        if not self._last_operations:
            messagebox.showinfo(
                "Sem dados",
                "Nenhuma apuração foi realizada ainda.",
            )
            return

        output_path = filedialog.asksaveasfilename(
            title="Salvar Excel de Debug",
            defaultextension=".xlsx",
            filetypes=[("Arquivo Excel", "*.xlsx")],
            initialfile=f"Memória de Cálculo - {self.investor_name.get().strip()}.xlsx",
        )
        if not output_path:
            return

        try:
            export_debug_excel(self._last_operations, output_path)
            messagebox.showinfo(
                "Exportação concluída",
                f"Excel de debug exportado com sucesso:\n{output_path}",
            )
        except Exception as exc:
            messagebox.showerror("Erro ao exportar debug", str(exc))
            traceback.print_exc()

    # ── Execução principal da apuração ────────────────────────────────────────

    def _run_apuracao(self):
        """Orquestra todo o fluxo: leitura → tipo de ativo → classificação → cálculo → Excel."""
        name        = self.investor_name.get().strip()
        cpf         = self.cpf.get().strip()
        folder      = self.selected_month_folder.get().strip()
        excel_path  = self.selected_excel_path.get().strip()
        cpf_limpo   = re.sub(r"[^\d]", "", cpf)

        if not name or not cpf_limpo or not folder or not excel_path:
            messagebox.showwarning("Campos pendentes", "Preencha todos os campos antes de calcular.")
            return

        if not os.path.exists(excel_path):
            messagebox.showwarning("Excel inválido", "Selecione um arquivo de planilha Excel válido.")
            return

        if is_file_locked_for_write(excel_path):
            messagebox.showerror(
                "Arquivo Excel Aberto",
                f"Feche o arquivo Excel antes de calcular:\n{os.path.basename(excel_path)}",
            )
            self.status_label.config(text="Feche o arquivo Excel antes de calcular.")
            return

        try:
            if not validar_cpf(cpf_limpo):
                messagebox.showwarning("CPF inválido", "O CPF informado é inválido.")
                return

            parts = os.path.normpath(folder).split(os.sep)
            if len(parts) < 2:
                raise ValueError(
                    "Caminho de pasta inválido. Estrutura esperada: .../<ANO>/<MM.AA>."
                )

            month_part = parts[-1]
            year_part  = parts[-2]

            month_match = re.match(r"^\s*(\d{1,2})(?:[./](\d{2}|\d{4}))?\s*$", month_part)
            if not month_match:
                raise ValueError(
                    f"O nome da pasta do mês ('{month_part}') é inválido. "
                    "Use o padrão 'MM.AA' ou 'MM.AAAA'."
                )

            year_digits = re.sub(r"[^0-9]", "", year_part)
            if year_digits:
                year = int(year_digits)
            elif month_match.group(2):
                year_text = month_match.group(2)
                year = 2000 + int(year_text) if len(year_text) == 2 else int(year_text)
            else:
                raise ValueError(
                    "Não foi possível identificar o ano da apuração pelo caminho da pasta. "
                    "Use uma estrutura como '.../2025/06.25' ou selecione uma pasta chamada '06.25'."
                )

            expected_month = int(month_match.group(1))
            if expected_month < 1 or expected_month > 12:
                raise ValueError(f"O mês '{expected_month:02d}' é inválido.")

            year_sheet_name = str(year)

            workbook   = openpyxl.load_workbook(excel_path)
            sheet_name = resolve_month_sheet_name(workbook, expected_month=expected_month, expected_year=year)

            if year_sheet_name not in workbook.sheetnames:
                raise RuntimeError(
                    f"A aba anual '{year_sheet_name}' não foi encontrada na planilha."
                )

            # 1. Lê as notas (Camada 1 → Camada 2)
            self.status_label.config(text="Lendo notas de corretagem...")
            operations = parse_month_folder(
                month_folder=folder,
                expected_year=year,
                expected_month=expected_month,
                expected_cpf_digits=cpf_limpo,
                password_map=self.password_map,
                ask_password_cb=self._pedir_senha_pdf,
                ask_quantity_cb=self._pedir_quantidade_trade,
                ask_manual_review_cb=self._revisao_manual_nota,
            )

            # 2. Resolve tickers ausentes
            for op in operations:
                if op.ticker and str(op.ticker).strip():
                    op.ticker = normalize_b3_ticker(op.ticker)
                    continue
                key = op.name
                if key in self.ticker_map:
                    op.ticker = normalize_b3_ticker(self.ticker_map[key])
                    continue
                t = self._pedir_ticker(op.note_file, op.ref_date, op.name)
                if not t:
                    raise RuntimeError(
                        f"Apuração cancelada: ticker não informado para '{op.name}'."
                    )
                t = normalize_b3_ticker(t) or ""
                self.ticker_map[key] = t
                self._save_ticker_map()
                op.ticker = t

            # 3. Detecta / solicita o tipo de ativo de cada operação
            self.status_label.config(text="Identificando tipo dos ativos...")
            self.resolve_asset_types(operations)

            # 4. Classifica (day / swing / fii)
            classify_operations(operations)

            # Guarda para exportação de debug
            self._last_operations = operations
            self.btn_debug.config(state="normal")

            # 5. Lê carteira anterior
            self.status_label.config(text="Lendo carteira do mês anterior...")
            initial_positions = read_initial_portfolio_from_year_sheet(
                workbook, year_sheet_name=year_sheet_name
            )

            # 6. Calcula resultados (agora com separação por tipo)
            self.status_label.config(text="Calculando resultados e carteira...")
            (
                result_swing_on_pn, result_swing_outros,
                result_day, result_fii,
                irrf_swing, irrf_day, irrf_fii,
                vendas_on_pn, vendas_outros,
                updated_positions, has_negative_positions,
            ) = process_results(operations, initial_positions)

            # 7. Atualiza planilha Excel
            self.status_label.config(text="Atualizando planilha Excel...")
            update_month_sheet(
                workbook,
                sheet_name=sheet_name,
                name=name,
                cpf=cpf_limpo,
                result_swing_on_pn=result_swing_on_pn,
                result_swing_outros=result_swing_outros,
                result_day=result_day,
                result_fii=result_fii,
                irrf_swing=irrf_swing,
                irrf_day=irrf_day,
                irrf_fii=irrf_fii,
                vendas_on_pn=vendas_on_pn,
                vendas_outros=vendas_outros,
            )

            write_portfolio_sheet_for_month(
                workbook, month_sheet_name=sheet_name, positions=updated_positions
            )
            workbook.save(excel_path)

            if has_negative_positions:
                messagebox.showwarning(
                    "Atenção - Carteira com quantidade negativa",
                    "Há quantidades negativas na carteira calculada.\n"
                    "Verifique se todas as notas foram importadas corretamente.",
                )

            self.status_label.config(
                text=(
                    f"Apuração de {sheet_name} concluída\n"
                    f"\n"
                    f"Resultado Swing Trade: {fmt(result_swing_on_pn + result_swing_outros)}\n"
                    f"Resultado Day Trade: {fmt(result_day)}\n"
                    f"Resultado FII: {fmt(result_fii)}"
                )
            )

        except Exception as exc:
            messagebox.showerror("Erro na apuração", str(exc))
            traceback.print_exc()
            self.status_label.config(text="Ocorreu um erro durante a apuração.")


###############################################################################
# MAIN: inicializa a GUI e executa o loop principal do Tkinter
###############################################################################

if __name__ == "__main__":
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    root = tk.Tk()
    app = ApuracaoB3App(root, base_dir=BASE_DIR)
    root.mainloop()
